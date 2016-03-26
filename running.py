#!/usr/bin/env python3

# References
# http://nedbatchelder.com/text/iter/iter.html#31

import openpyxl
import datetime
import calendar
# import time
from datetime import timedelta
# from datetime import datetime

from flask import Flask, render_template, request, g
import sqlite3
import sql

import matplotlib.pyplot as plt

DATABASE = '/Users/sroberts/Dropbox/TMT/Python/SQL/races.sqlite'

app = Flask(__name__)

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
    return db

def time2timedelta(t):
    td = timedelta(hours=t.hour,minutes=t.minute,seconds=t.second)
    return td

def iso_year_start(iso_year):
    "The gregorian calendar date of the first day of the given ISO year"
    fourth_jan = datetime.date(iso_year, 1, 4)
    delta = datetime.timedelta(fourth_jan.isoweekday()-1)
    return fourth_jan - delta 
    
def iso_to_gregorian(iso_year, iso_week, iso_day):
    "Gregorian calendar date for the given ISO year, week and day"
    year_start = iso_year_start(iso_year)
    return year_start + datetime.timedelta(days=iso_day-1, weeks=iso_week-1)
    
def greg_YMDstr_to_datetime(greg_str):
    # Converts string in YYYY-MM-DD format into datetime object
    return(datetime.datetime.strptime(greg_str,'%Y-%m-%d')

def iso_YWDstr_to_datetime(iso_str):
    # Converts string in YYYY-WXX-D format into a datetime object
    return(datetime.datetime.strptime(greg_str,'%Y-W%W-%w')
    
def read_spreadsheet( sheet ):

    # define indices for spreadsheet row and dictionary row
    ssrow = 2
    dicrow = 1
    myRun = {}

    while sheet.cell(row=ssrow, column=1).value:
    
        myRun[dicrow] = {}
    
        # assign columns from spreadsheet to variables
        myRun[dicrow]['date'] = sheet.cell(row = ssrow, column = 1).value
        myRun[dicrow]['weekday'] = sheet.cell(row = ssrow, column = 2).value
        myRun[dicrow]['description'] = sheet.cell(row = ssrow, column = 3).value
        myRun[dicrow]['objective'] = sheet.cell(row = ssrow, column = 4).value
        myRun[dicrow]['notes'] = sheet.cell(row = ssrow, column = 5).value
        myRun[dicrow]['distance'] = sheet.cell(row = ssrow, column = 6).value
        myRun[dicrow]['time_run'] = sheet.cell(row = ssrow, column = 7).value
        myRun[dicrow]['ave_speed'] = sheet.cell(row = ssrow, column = 8).value
        myRun[dicrow]['time_recovery'] = sheet.cell(row = ssrow, column = 9).value
        myRun[dicrow]['time_easy'] = sheet.cell(row = ssrow, column = 10).value
        myRun[dicrow]['time_tempo'] = sheet.cell(row = ssrow, column = 11).value
        myRun[dicrow]['time_interval'] = sheet.cell(row = ssrow, column = 12).value
        myRun[dicrow]['time_repetition'] = sheet.cell(row = ssrow, column = 13).value
        myRun[dicrow]['percent_low'] = sheet.cell(row = ssrow, column = 14).value
        myRun[dicrow]['percent_med'] = sheet.cell(row = ssrow, column = 15).value
        myRun[dicrow]['percent_high'] =  sheet.cell(row = ssrow, column = 16).value
        myRun[dicrow]['time_80_20'] = sheet.cell(row = ssrow, column = 17).value
        myRun[dicrow]['shoes'] = sheet.cell(row = ssrow, column = 18).value
        myRun[dicrow]['dist_sum_all'] = sheet.cell(row = ssrow, column = 22).value
        myRun[dicrow]['dist_monthly'] = sheet.cell(row = ssrow, column = 23).value
        myRun[dicrow]['ave_dist_month'] = sheet.cell(row = ssrow, column = 24).value 
 
        ssrow += 1
        dicrow += 1
        
    return myRun
    
def calc_data( myRun ):
    for k,v in myRun.items():
        # since myRun is a dictionary of dictionaries k is the integer
        # key dicrow and v is the dictionary of items for that entry
        
        # .isocalendar() returns a 3-tuple with (year, wk num, wk day)

#         print(k)
        myRun[k]['isodate'] = myRun[k]['date'].isocalendar()
        myRun[k]['time_run'] = time2timedelta(datetime.datetime.strptime(myRun[k]['time_run'],'%H:%M:%S'))

        if myRun[k]['time_recovery'] != None:
            myRun[k]['time_recovery'] = time2timedelta(datetime.datetime.strptime(myRun[k]['time_recovery'],'%H:%M:%S'))
        
        if myRun[k]['time_easy'] != None:
            myRun[k]['time_easy'] = time2timedelta(datetime.datetime.strptime(myRun[k]['time_easy'],'%H:%M:%S'))
            
        if myRun[k]['time_tempo'] != None:
            myRun[k]['time_tempo'] = time2timedelta(datetime.datetime.strptime(myRun[k]['time_tempo'],'%H:%M:%S'))
            
        if myRun[k]['time_interval'] != None:
            myRun[k]['time_interval'] = time2timedelta(datetime.datetime.strptime(myRun[k]['time_interval'],'%H:%M:%S'))
            
        if myRun[k]['time_repetition'] != None:
            myRun[k]['time_repetition'] = time2timedelta(datetime.datetime.strptime(myRun[k]['time_repetition'],'%H:%M:%S'))
            
    return(myRun)
        

def get_data_for_rowrange(myRun, id, date_s, date_e, row_s, row_e):   
    # Fill in period data
    # id is a text string identifying the period
    # date_s and date_e are datetime objects of the start and end of the period
    # row_s and row_e are the starting and ending rows within the period
    # function returns a dictionary data structure
    
    period = {}
    period['id'] = id
    period['start_date'] = date_s
    period['end_date'] = date_e
    period['start_row_id'] = row_s
    period['end_row_id'] = row_e
    
#     print('get_data_for_rowrange', id, date_s, date_e, row_s, row_e)
    
    # Add up the mileage and running time for the period
    period['kms'] = 0
    period['time'] = timedelta(hours = 0)

    entries = []
    for i in range(row_s, row_e+1):
#         print('i = ',i,'   date = ', myRun[i]['date'])
        period['kms'] += myRun[i]['distance']
        period['time'] += myRun[i]['time_run']
        entries.append([myRun[i]['date'],myRun[i]['distance'],myRun[i]['time_run']])
        
    period['Entries'] = entries
        
    return period

def print_periods( id, periods ):
    for k,v in periods.items():
        print(
                id,'%02d' % k,v['id'],':',
                v['start_date'].strftime('%Y-%m-%d'),
                v['end_date'].strftime('%Y-%m-%d'),
                'kms:', '% 6.1f' % v['kms'],
                'time:', v['time']
              )
       
    x = []
    y = []       
    for k,v in periods.items():
#         x.append(k)
        x.append(v['end_date'])
        y.append(v['kms'])
#         plt.plot(x, y, 'ro')
#         plt.bar(x, y, 0.35, color='r')
        plt.plot(x, y)
    plt.show()


def calc_averages(periods):
    # find highest id
    high_id = 0
    for k,v in periods.items():
        if k > high_id:
            high_id = k
    
    count = 0
    total = 0
    for idx in range(high_id-1, high_id-65, -1):
        count += 1
        total += periods[idx]['kms']
#         print('total = ', total)
#         print('count = ', count)
        ave = total/count
#         print('ave = %3.1f' % ave)
#         print(periods[idx])
        if count in [1,2,4,8,16,32,64]:
            print('ave for last % 2d' % count, 'weeks = % 3.1f' % ave)
    

def get_weeks(myRun):
    start_ssrow = 1
    
    week = myRun[start_ssrow]['isodate'][1] # the iso date week
    year = myRun[start_ssrow]['isodate'][0] # the iso date year
    weekstats = {}
    idx = 0
    
    for k,v in myRun.items():
#         print('list of myRun: k = ',k,' date = ', v['date'])
        if v['isodate'][1] != week:     # we found the end entry for the week
            end_ssrow = k-1
            date_s = iso_to_gregorian(year, week, 1)    # Monday
            date_e = iso_to_gregorian(year, week, 7)    # Sunday
            idx += 1
            weekstats[idx] = {}
            sw = str(week)
            if len(sw) == 1:
                sw = '0'+sw
            weekstats[idx] = get_data_for_rowrange(myRun, str(year)+'-'+sw, 
                date_s, date_e, start_ssrow, end_ssrow)
            start_ssrow = k
            week = myRun[start_ssrow]['isodate'][1] # the iso date week
            year = myRun[start_ssrow]['isodate'][0] # the iso date year
            
    # end the last week
    end_ssrow = k
    date_s = iso_to_gregorian(year, week, 1)
    date_e = iso_to_gregorian(year, week, 7)
    idx += 1
    weekstats[idx] = {}
    sw = str(week).rjust(2,'0')
    
    weekstats[idx] = get_data_for_rowrange(myRun, str(year)+'-'+sw, 
        date_s, date_e, start_ssrow, end_ssrow)
        
    return weekstats
    
    
def get_months(myRun):
    start_ssrow = 1
    monthstats = {}   
    
    # Get the month and year
    d = myRun[start_ssrow]['date']
    month = d.month
    year = d.year
    sm = str(month)
    if len(sm) == 1:
        sm = '0'+sm
    
    # Get start day of week and number of days in month
    [sd, mm] = calendar.monthrange(year, month)
    
    idx = 0
    
    for k,v in myRun.items():
        if v['date'].month != month:
            # The final entry for this period is the previous one
            end_ssrow = k-1
            
            date_s = datetime.datetime(year, month, 1)
            date_e = datetime.datetime(year, month, mm)
            idx += 1
            monthstats[idx] = {}

            monthstats[idx] = get_data_for_rowrange(myRun, str(year)+'-'+sm, 
                date_s, date_e, start_ssrow, end_ssrow)
                
            start_ssrow = k
            d = myRun[start_ssrow]['date']
            month = d.month
            year = d.year
            sm = str(month)
            if len(sm) == 1:
                sm = '0'+sm
            [sd, mm] = calendar.monthrange(year, month)
 
    end_ssrow = k
    
    date_s = datetime.datetime(year, month, 1)
    date_e = datetime.datetime(year, month, mm)
    
    idx += 1
    monthstats[idx] = {}
    
    sm = str(month)
    if len(sm) == 1:
        sm = '0'+sm
    
    monthstats[idx] = get_data_for_rowrange(myRun, str(year)+'-'+sm, 
        date_s, date_e, start_ssrow, end_ssrow)     
              
    return monthstats
    
    
def calc_8020(weekstats, myRun):
    # Calculate 80/20 training intensity over the weeks that have entries
    # Jack Daniels intensity points
    # Recovery - 0.15 points/minute (my guess, not in JD's book)
    # Easy - 0.2 points/minute
    # Threshold - 0.6 points/minute
    # 10k pace - 0.8 points/minute
    # Interval - 1.0 points/minute
    # Repetition - 1.5 points/minute
    
    zt = timedelta(0)
    
    
    for k,v in weekstats.items():
        recovery_time = zt
        easy_time = zt
        tempo_time = zt
        interval_time = zt
        repetition_time = zt
        
        for i in range(weekstats[k]['start_row_id'],weekstats[k]['end_row_id']+1):
            if myRun[i]['time_recovery'] != None:
                recovery_time += myRun[i]['time_recovery']
            if myRun[i]['time_easy'] != None:
                easy_time += myRun[i]['time_easy']
            if myRun[i]['time_tempo'] != None:                
                tempo_time += myRun[i]['time_tempo']
            if myRun[i]['time_interval'] != None:
                interval_time += myRun[i]['time_interval']
            if myRun[i]['time_repetition'] != None:
                repetition_time += myRun[i]['time_repetition']
                
        # Check that there is at least one non-zero time in the data
        if recovery_time != zt or easy_time != zt or tempo_time != zt or interval_time != zt or repetition_time != zt:     

            recovery_int = recovery_time.total_seconds()/60 * 0.15
            easy_int = easy_time.total_seconds()/60 * 0.2
            tempo_int = tempo_time.total_seconds()/60 * 0.6
            interval_int = interval_time.total_seconds()/60 * 1.0
            repetition_int = repetition_time.total_seconds()/60 * 1.5
            jd_intensity = recovery_int + easy_int + tempo_int + interval_int + repetition_int

            print('Week', k, 'Starting', weekstats[k]['start_date'], 
                'Rec:', recovery_time, 'E:', easy_time, 
                'T:', tempo_time, 'I:', interval_time, 
                'Rep:', repetition_time)
                              
            sum_low = easy_time + recovery_time
            sum_midhigh = tempo_time + interval_time + repetition_time
            sum_total = sum_low + sum_midhigh
            ssum_low = sum_low.total_seconds()
            ssum_midhigh = sum_midhigh.total_seconds()
            
            p8020 = ssum_low/(ssum_low+ssum_midhigh)*100.0
            print('Time Total:',sum_total,'Percent 80/20:', '%.1f' % p8020,'%',
                'Intensity:','%.1f' % jd_intensity)    
    

             
def add_to_shoes( shoes, shoe, date, kms ):
    if shoe not in shoes:
        shoes[shoe] = [date, date, kms]
    else:
        shoes[shoe][1] = date
        shoes[shoe][2] += kms
    return(shoes)
    
def get_shoe_mileage(shoes, myRun):
    for k,v in myRun.items():
        shoes = add_to_shoes(shoes,v['shoes'],v['date'],v['distance'])
    return shoes
    
  
# Main Program  
def main():
    path = r'/Users/sroberts/Google Drive/Exercise/ExerciseLog.xlsx'  
    # Open the spreadsheet
    wb=openpyxl.load_workbook( path, data_only=True )

    # read the running log by sheet name
    sheet = wb.get_sheet_by_name('Running Log')

    # Read in the spreadsheet to a dictionary structure
    # myRun = read_spreadsheet( sheet )

#     conn = get_db()
#     cur = conn.cursor()
    [conn, cur] = sql.load_database(DATABASE)
    myRun = sql.get_running_dict(conn, cur)

    # Calculate additional data
    calc_data( myRun )

    # Calculate and print data for weeks
    print('\nWeekly Run Stats')
    weekstats = get_weeks(myRun)
    print_periods('Week:', weekstats)
    calc_averages(weekstats)

    # Calculate and print data for months
    print('\nMonthly Run Stats')
    monthstats = get_months(myRun)
    print_periods('Month:', monthstats)

    # Calculate 80/20 running
    print('\nPercent 80/20 Running:')
    calc_8020(weekstats, myRun)

    # Calculate and print data for shoes
    shoes = {}
    shoes = get_shoe_mileage(shoes, myRun)

    print('\nshoe usage sorted by last date used:')
    for k,v in sorted(shoes.items(), key=lambda x: x[1][1], reverse = True):
        # lambda is sorting on last use x[1] is the list, and [1] in list is last use date
        print('[first use:', v[0].strftime('%Y-%m-%d'), 'last use:', 
            v[1].strftime('%Y-%m-%d'), 'kms:', '%.1f' % v[2], ']', k )
    
   
if __name__ == "__main__":
	main()
