#!/usr/bin/env python3

import re
import csv
import sqlite3
import time
import logging
from dataclasses import dataclass
from datetime import datetime, timedelta
from typing import Any
import numpy as np

logger = logging.getLogger(__name__)
from markupsafe import Markup, escape
from urllib.parse import quote as urlquote

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import BLUE

import matplotlib
# See https://stackoverflow.com/questions/53684971 for why Agg backend is required with Flask
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as ticker

font_url_style = Font(color=BLUE, underline='single')
bold_style = Font(bold=True)
align_hv_cen_style = Alignment(horizontal='center', vertical='center')
align_ver_cen_style = Alignment(vertical='center')
align_wrap_cen_style = Alignment(wrap_text=True, vertical='center')


def parse_datestr(date_str: str) -> list[str]:
    """Split a YYYY-MM-DD (or ISO week) string into [year, month_or_week, day]."""
    m = re.search(r'(^\d*)-(\d*)-(\d*)', date_str)
    if not m:
        raise ValueError(f'Cannot parse date string: {date_str!r}')
    return [m.group(1), m.group(2), m.group(3)]


def scrub_timestr(time_str: str) -> str:
    """Normalise a time string to H:MM:SS format."""
    m = re.search(r'(^\d*):(\d*):(\d*)', time_str)
    hours = str(int(m.group(1)))
    mins = str(int(m.group(2)))
    secs_val = str(int(m.group(3)))
    return f'{hours}:{mins.zfill(2)}:{secs_val.zfill(2)}'


def timestr2pacestr(time_str: str) -> str:
    """Convert H:MM:SS time string to MM:SS pace string."""
    m = re.search(r'(^\d*):(\d*):(\d*)', time_str)
    mins = str(int(m.group(2)))
    secs_val = str(int(m.group(3)))
    return f'{mins}:{secs_val.zfill(2)}'


def scrub_pace(pace_str: str) -> str:
    """Normalise a MM:SS pace string."""
    m = re.search(r'(^\d*):(\d*)', pace_str)
    mins = str(int(m.group(1)))
    secs_val = str(int(m.group(2)))
    return f'{mins}:{secs_val.zfill(2)}'


def get_running_dict(conn, cur) -> dict:
    """Return all workout log entries as a dict keyed by sequential index."""
    cur.execute(
        'SELECT Log.id, date, location, objective, notes, dist, time, pace, recovery, easy, '
        'threshold, interval, repetition, p8020, jd_int, shortName FROM Log '
        'JOIN Shoes ON Log.ShoeID = Shoes.id '
        'ORDER BY date'
    )
    my_run = {}
    for idx, row in enumerate(cur, start=1):
        date = datetime.strptime(row[1], '%Y-%m-%d')
        my_run[idx] = {
            'id': row[0],
            'date': date,
            'weekday': date.strftime('%A'),
            'description': row[2],
            'objective': row[3],
            'notes': row[4],
            'distance': row[5],
            'time_run': row[6],
            'pace': row[7],
            'time_recovery': row[8],
            'time_easy': row[9],
            'time_tempo': row[10],
            'time_interval': row[11],
            'time_repetition': row[12],
            'time_80_20': row[13],
            'JD_Intensity': row[14],
            'shoes': row[15],
        }
    return my_run


def read_ss(sheet, conn, cur) -> None:
    """Import workout rows from an openpyxl worksheet starting at row 5."""
    ss_row = 5

    def cell_or_default(row, col, default):
        val = sheet.cell(row=row, column=col).value
        return default if not val else str(val)

    while sheet.cell(row=ss_row, column=1).value:
        logger.debug('date = %s', sheet.cell(row=ss_row, column=1).value)

        date = sheet.cell(row=ss_row, column=1).value
        location = sheet.cell(row=ss_row, column=2).value
        objective = sheet.cell(row=ss_row, column=3).value
        notes = sheet.cell(row=ss_row, column=4).value
        distance = sheet.cell(row=ss_row, column=5).value
        time_val = cell_or_default(ss_row, 6, '0:00:00')
        recovery = cell_or_default(ss_row, 8, '0:00:00')
        easy = cell_or_default(ss_row, 9, '0:00:00')
        threshold = cell_or_default(ss_row, 10, '0:00:00')
        interval = cell_or_default(ss_row, 11, '0:00:00')
        repetition = cell_or_default(ss_row, 12, '0:00:00')

        if recovery == '0:00:00' and easy == '0:00:00' and threshold == '0:00:00' and interval == '0:00:00' and repetition == '0:00:00':
            easy = time_val

        shoes = sheet.cell(row=ss_row, column=15).value
        cur.execute('SELECT id FROM shoes WHERE shortName = ?', (shoes,))
        shoe_id = cur.fetchone()[0]

        cur.execute('SELECT id FROM wo_type ORDER BY id LIMIT 1')
        result = cur.fetchone()
        wo_type_id = result[0] if result else 1

        logger.debug('times %s %s %s %s %s %s', time_val, recovery, easy, threshold, interval, repetition)
        add_workout(conn, cur, date, location, wo_type_id, objective, notes, distance, recovery, easy, threshold, interval, repetition, shoe_id)
        ss_row += 1


def write_ss(conn, cur, xl_file: str) -> None:
    """Write workout log to an Excel spreadsheet."""
    wb = openpyxl.Workbook()
    ws = wb.worksheets[0]

    row_num = 4
    col = 1

    headers = ['Date', 'Location', 'Objective', 'Notes', 'Dist', 'Time',
               'Pace', 'Recovery', 'Easy', 'Threshold', 'Interval',
               'Repetition', '80/20', 'Intensity', 'Shoes']

    for col, header in enumerate(headers, start=1):
        ws.cell(row=row_num, column=col).value = header
        ws.cell(row=row_num, column=col).alignment = align_hv_cen_style
        ws.cell(row=row_num, column=col).font = bold_style

    col_alignments = [
        align_hv_cen_style, align_wrap_cen_style, align_wrap_cen_style, align_wrap_cen_style,
        align_hv_cen_style, align_hv_cen_style, align_hv_cen_style, align_hv_cen_style,
        align_hv_cen_style, align_hv_cen_style, align_hv_cen_style, align_hv_cen_style,
        align_hv_cen_style, align_hv_cen_style, align_wrap_cen_style,
    ]

    intro, thead, tbody, summary = get_workouts(cur)
    ss_row = row_num

    for row in tbody:
        ss_row += 1
        for col, (value, alignment) in enumerate(zip(row, col_alignments), start=1):
            ws.cell(row=ss_row, column=col).value = value
            ws.cell(row=ss_row, column=col).alignment = alignment

    col_widths = {'A': 8.0, 'B': 20.0, 'C': 10.0, 'D': 40.0, 'E': 10.0,
                  'F': 10.0, 'G': 10.0, 'H': 10.0, 'I': 10.0, 'J': 10.0,
                  'K': 10.0, 'L': 10.0, 'M': 10.0, 'N': 10.0, 'O': 20.0}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(xl_file)


def secs(time_str: str) -> float:
    """Return total seconds in a H:MM:SS time string."""
    x = time.strptime(time_str, '%H:%M:%S')
    return timedelta(hours=x.tm_hour, minutes=x.tm_min, seconds=x.tm_sec).total_seconds()


def str_time(total_seconds: float) -> str:
    """Return H:MM:SS string for a number of seconds."""
    return str(timedelta(seconds=round(total_seconds)))


def load_csv_health(conn, cur, csv_filename: str) -> None:
    """Import health records from a CSV file."""
    try:
        with open(csv_filename) as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                try:
                    add_health(conn, cur, row['Date'], row['Weight (kg)'], row['Waist'],
                               row['Waist at bb'], row['Hip'], row['Chest'], row['Notes'], row['HR'])
                except Exception as e:
                    logger.error('Error importing CSV file for Health: %s', e)
                    raise
        conn.commit()
        logger.info('Health CSV import succeeded')
    except Exception:
        logger.error('Health CSV import failed')


def add_health(conn, cur, date: str, weight, small_waist, bb_waist, hip, chest, notes, hr) -> None:
    """Insert a health record."""
    try:
        cur.execute(
            'INSERT INTO Health (date, weight, smallWaist, bbWaist, hip, chest, notes, HR) '
            'VALUES (?,?,?,?,?,?,?,?)',
            (date, weight, small_waist, bb_waist, hip, chest, notes, hr)
        )
    except Exception as e:
        logger.error('Error - unable to enter health, date=%s: %s', date, e)
    conn.commit()


def change_health(conn, cur, id, date: str, weight, small_waist, bb_waist, hip, chest, notes, hr) -> None:
    """Update an existing health record."""
    try:
        cur.execute(
            'UPDATE Health SET date=?, weight=?, smallWaist=?, bbWaist=?, hip=?, chest=?, notes=?, HR=? '
            'WHERE id = ?',
            (date, weight, small_waist, bb_waist, hip, chest, notes, hr, id)
        )
    except Exception as e:
        logger.error('Error - unable to update health record, id=%s: %s', id, e)
    conn.commit()


def get_health(cur, id) -> tuple:
    """Return a single health record by id."""
    cur.execute(
        'SELECT id, date, weight, smallWaist, bbWaist, hip, chest, HR, notes FROM Health WHERE id = ?',
        (id,)
    )
    return cur.fetchone()


def get_health_list(cur) -> tuple[list, list, list, str]:
    """Return intro/thead/tbody/summary for all health records."""
    cur.execute('SELECT id, date, weight, smallWaist, bbWaist, hip, chest, HR, notes FROM Health ORDER BY date DESC')
    intro = ['Listing of Health']
    thead = ['ID', 'Date', 'Weight', 'Waist', 'Waist at BB', 'Hip', 'Chest', 'HR', 'Notes']
    tbody = []
    for row in cur:
        row = list(row)
        row[0] = Markup(f'<strong><a href=/ChangeHealth/{row[0]}>{row[0]}</a></strong>')
        tbody.append(row)
    summary = f'Total of {len(tbody)} Entries'
    return intro, thead, tbody, summary


def get_weight_report(cur) -> None:
    """Generate and save a weight trend plot to static/weight.png."""
    cur.execute('SELECT date FROM Health WHERE (weight > 0) ORDER BY date DESC')
    dates = []
    for count, row in enumerate(cur):
        if count >= 200:
            break
        dates.append(row[0])

    datelist = []
    weightlist = []
    for str_date in dates:
        date = datetime.strptime(str_date, '%Y-%m-%d')
        datelist.append(date)
        date_3_days_ago = date - timedelta(days=3)
        date_3_days_future = date + timedelta(days=3)
        start = date_3_days_ago.strftime('%Y-%m-%d')
        end = date_3_days_future.strftime('%Y-%m-%d')
        try:
            cur.execute(
                'SELECT AVG(weight) FROM Health WHERE (date between ? AND ?) AND weight > 0.0',
                (start, end)
            )
            weightlist.append(cur.fetchone())
        except Exception as e:
            logger.error('Weight report query error: %s', e)

    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
    plt.gca().xaxis.set_major_locator(ticker.MultipleLocator(90))
    plt.plot(datelist, weightlist)
    plt.gcf().autofmt_xdate()
    plt.savefig('static/weight.png')
    plt.clf()


def retire_shoe(conn, cur, shoe_id) -> None:
    """Toggle the retired flag for a shoe."""
    cur.execute('SELECT retired FROM Shoes WHERE id = ?', (shoe_id,))
    row = cur.fetchone()
    if row:
        cur.execute('UPDATE Shoes SET retired = ? WHERE id = ?', (0 if row[0] else 1, shoe_id))
        conn.commit()


def get_all_shoes(cur) -> tuple[list, list, list, str]:
    """Return intro/thead/tbody/summary for all shoes with retire/unretire links."""
    cur.execute(
        'SELECT Shoes.id, Shoes.shortName, Shoes.longName, Shoes.retired, '
        'COALESCE(ROUND(SUM(Log.dist), 1), 0), COALESCE(MAX(Log.date), "never") '
        'FROM Shoes LEFT JOIN Log ON Log.shoeID = Shoes.id '
        'GROUP BY Shoes.id ORDER BY Shoes.retired, Shoes.shortName'
    )
    intro = ["Click Retire or Unretire to change a shoe's status"]
    thead = ['Short Name', 'Long Name', 'Status', 'Distance (km)', 'Last Used', 'Action']
    tbody = []
    for shoe_id, short, long_name, retired, dist, last in cur:
        status = 'Retired' if retired else 'Active'
        label = 'Unretire' if retired else 'Retire'
        action = Markup(f'<a href="/RetireShoe/{shoe_id}">{label}</a>')
        tbody.append([escape(short), escape(long_name), status, dist, last, action])
    summary = f'Total of {len(tbody)} shoes'
    return intro, thead, tbody, summary


def add_shoes(conn, cur, short_name: str, long_name: str) -> None:
    """Insert a new shoe record."""
    logger.info('Adding new shoes: %s / %s', short_name, long_name)
    try:
        cur.execute(
            'INSERT INTO Shoes (shortName, longName, retired) VALUES (?,?,?)',
            (short_name, long_name, 0)
        )
    except Exception as e:
        logger.error('Error adding shoe (may already exist): %s', e)
    conn.commit()


def get_shoes(cur) -> tuple[list, list, list, str]:
    """Return intro/thead/tbody/summary for shoes that have logged workouts."""
    cur.execute(
        'SELECT Shoes.id, shortName, longName, sum(dist), max(date) FROM Shoes '
        'JOIN Log WHERE log.shoeID = Shoes.id GROUP BY log.shoeID ORDER BY max(date) DESC'
    )
    intro = ['Listing of Shoes']
    thead = ['ID', 'Short Name', 'Long Name', 'Dist', 'Date Last Used']
    tbody = []
    for row in cur:
        row = list(row)
        row[3] = f'{row[3]:.1f}'
        tbody.append(row)
    summary = f'Total of {len(tbody)} Entries'
    return intro, thead, tbody, summary


@dataclass
class WorkoutData:
    isodate_str: str
    total_time_str: str
    total_time: float
    pace: str
    secs_pace: float
    secs_recovery: float
    secs_easy: float
    secs_threshold: float
    secs_interval: float
    secs_repetition: float
    p8020: float
    jd_int: float


def get_workout_calculated_data(date: str, distance, recovery: str, easy: str,
                                threshold: str, interval: str, repetition: str) -> WorkoutData:
    """Compute derived workout metrics from raw inputs."""
    secs_recovery = secs(recovery)
    secs_easy = secs(easy)
    secs_threshold = secs(threshold)
    secs_interval = secs(interval)
    secs_repetition = secs(repetition)

    total_time = secs_recovery + secs_easy + secs_threshold + secs_interval + secs_repetition
    total_time_str = str_time(total_time)
    secs_pace = (total_time / float(distance)) if float(distance) > 0.0 else 0.0
    pace = str_time(secs_pace)

    p8020 = (100.0 * (secs_recovery + secs_easy) / total_time) if total_time > 0 else 0.0

    jd_int = (secs_recovery / 60.0 * 0.15 +
              secs_easy / 60.0 * 0.2 +
              secs_threshold / 60.0 * 0.6 +
              secs_interval / 60.0 * 1.0 +
              secs_repetition / 60.0 * 1.5)

    iso = datetime.strptime(date, '%Y-%m-%d').isocalendar()
    isodate_str = f'{iso[0]}-{str(iso[1]).zfill(2)}-{iso[2]}'

    return WorkoutData(
        isodate_str=isodate_str,
        total_time_str=total_time_str,
        total_time=total_time,
        pace=pace,
        secs_pace=secs_pace,
        secs_recovery=secs_recovery,
        secs_easy=secs_easy,
        secs_threshold=secs_threshold,
        secs_interval=secs_interval,
        secs_repetition=secs_repetition,
        p8020=p8020,
        jd_int=jd_int,
    )


def change_workout(conn, cur, id, date: str, location: str, wo_type_id: int, objective: str,
                   notes: str, distance, recovery: str, easy: str, threshold: str,
                   interval: str, repetition: str, shoe_id: int) -> None:
    """Update an existing workout log entry."""
    logger.info('Changing workout ID: %s', id)
    wd = get_workout_calculated_data(date, distance, recovery, easy, threshold, interval, repetition)
    try:
        cur.execute(
            'UPDATE Log SET date=?, location=?, wo_type=?, objective=?, notes=?, dist=?, '
            'time=?, time_secs=?, pace=?, pace_secs=?, recovery=?, '
            'recovery_secs=?, easy=?, easy_secs=?, threshold=?, '
            'threshold_secs=?, interval=?, interval_secs=?, repetition=?, '
            'repetition_secs=?, p8020=?, jd_int=?, shoeID=?, isodate=? WHERE id=?',
            (date, location, wo_type_id, objective, notes, float(distance),
             wd.total_time_str, wd.total_time, wd.pace, wd.secs_pace,
             recovery, wd.secs_recovery, easy, wd.secs_easy,
             threshold, wd.secs_threshold, interval, wd.secs_interval,
             repetition, wd.secs_repetition, wd.p8020, wd.jd_int,
             shoe_id, wd.isodate_str, id)
        )
        conn.commit()
    except Exception as e:
        logger.error('Error updating workout id=%s: %s', id, e)


def add_workout(conn, cur, date: str, location: str, wo_type_id: int, objective: str,
                notes: str, distance, recovery: str, easy: str, threshold: str,
                interval: str, repetition: str, shoe_id: int) -> None:
    """Insert a new workout log entry."""
    logger.info('Adding workout for date: %s', date)
    wd = get_workout_calculated_data(date, distance, recovery, easy, threshold, interval, repetition)
    cur.execute(
        'INSERT INTO Log (date, location, wo_type, objective, notes, dist, time, time_secs, '
        'pace, pace_secs, recovery, recovery_secs, easy, easy_secs, threshold, threshold_secs, '
        'interval, interval_secs, repetition, repetition_secs, p8020, jd_int, shoeID, isodate) '
        'VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
        (date, location, wo_type_id, objective, notes, float(distance),
         wd.total_time_str, wd.total_time, wd.pace, wd.secs_pace,
         recovery, wd.secs_recovery, easy, wd.secs_easy,
         threshold, wd.secs_threshold, interval, wd.secs_interval,
         repetition, wd.secs_repetition, wd.p8020, wd.jd_int,
         shoe_id, wd.isodate_str)
    )
    conn.commit()


def get_workout(cur, id) -> tuple:
    """Return a single workout record by id."""
    cur.execute(
        'SELECT date, location, type, objective, notes, dist, recovery, easy, '
        'threshold, interval, repetition, shortName FROM Log '
        'JOIN Shoes ON Log.ShoeID = Shoes.id '
        'JOIN wo_type ON Log.wo_type = wo_type.id '
        'WHERE Log.id = ?',
        (id,)
    )
    return cur.fetchone()


def get_workouts(cur) -> tuple[list, list, list, str]:
    """Return intro/thead/tbody/summary for all workout log entries."""
    cur.execute(
        'SELECT Log.id, date, location, objective, notes, dist, time, pace, recovery, easy, '
        'threshold, interval, repetition, p8020, jd_int, shortName, type FROM Log '
        'JOIN Shoes ON Log.ShoeID = Shoes.id '
        'JOIN wo_type ON Log.wo_type = wo_type.id '
        'ORDER BY date DESC'
    )
    intro = ['Listing of Workouts']
    thead = ['ID', 'Run Date', 'Workout Description', 'Dist', 'Time', 'Pace', 'Time in Zones',
             '80/20', 'JD Intensity', 'Shoes']
    tbody = []
    for row in cur:
        row = list(row)
        id_tag = Markup(f'<strong><a href=/ChangeWorkout/{row[0]}>{row[0]}</a></strong>')
        date = Markup(str(row[1]))
        location = row[2] or ''
        objective = row[3] or ''
        notes = row[4] or ''
        dist = str(row[5])
        wo_type_text = row[16]

        description = (Markup('<strong>Location: </strong>') + escape(location) +
                       Markup('<br /><strong>WO Type: </strong>') + escape(wo_type_text) +
                       Markup('<br /><strong>Objective: </strong>') + escape(objective) +
                       Markup('<br /><strong>Notes: </strong>') + escape(notes))

        p8020 = f'{ row[13]: .1f}%'
        intensity = f'{ row[14]: .2f}'
        zones = zones_str(row[8], row[9], row[10], row[11], row[12])

        tbody.append([id_tag, date, description, dist, row[6], row[7], zones,
                      p8020, intensity, row[15]])

    summary = f'Total of {len(tbody)} Workouts'
    return intro, thead, tbody, summary


def zones_str(recovery: str, easy: str, threshold: str, interval: str, repetition: str) -> Markup:
    """Build an HTML summary of time spent in each training zone."""
    zones = Markup('')
    if recovery != '0:00:00':
        zones += Markup('<strong>Recovery: </strong>') + escape(recovery) + Markup('\n')
    if easy != '0:00:00':
        zones += Markup('<strong>Easy: </strong>') + escape(easy) + Markup('\n')
    if threshold != '0:00:00':
        zones += Markup('<strong>Threshold: </strong>') + escape(threshold) + Markup('\n')
    if interval != '0:00:00':
        zones += Markup('<strong>Interval: </strong>') + escape(interval) + Markup('\n')
    if repetition != '0:00:00':
        zones += Markup('<strong>Repetition: </strong>') + escape(repetition) + Markup('\n')
    return zones


def add_event(conn, cur, race_name: str, dist: float, min_elev: int, max_elev: int, gain_elev: int) -> None:
    """Insert a race event."""
    logger.info('Adding event: %s', race_name)
    try:
        cur.execute(
            'INSERT INTO Events (eventname, dist, min_elev, max_elev, gain_elev) VALUES (?,?,?,?,?)',
            (race_name, dist, min_elev, max_elev, gain_elev)
        )
    except Exception as e:
        logger.error('Event already exists or error: %s', e)
    conn.commit()


def test_load_result(conn, cur, csv_filename: str, event_name: str, date: str,
                     dist: float, min_elev: int, max_elev: int, gain_elev: int) -> None:
    """Log CSV rows for a race event without inserting them (diagnostic)."""
    logger.info('Test loading results for %s', event_name)
    cur.execute('SELECT id FROM Events WHERE (eventname = ?) LIMIT 1', (event_name,))
    event_id = cur.fetchone()[0]
    logger.debug('Found eventID %s for %s', event_id, event_name)
    with open(csv_filename) as csvfile:
        for row in csv.DictReader(csvfile):
            logger.debug('test row: %s %s %s', row.get('name'), row.get('time'), row.get('pace'))


def load_result(conn, cur, csv_filename: str, event_name: str, date: str,
                dist: float, min_elev: int, max_elev: int, gain_elev: int) -> None:
    """Load race results from a CSV file into the database."""
    logger.info('Loading results for %s', event_name)

    cur.execute('SELECT id FROM Events WHERE (eventname = ?) LIMIT 1', (event_name,))
    event_id = cur.fetchone()[0]

    try:
        cur.execute('INSERT INTO Races (eventID, date) VALUES (?,?)', (event_id, date))
    except Exception as e:
        logger.error('Race already exists or error: %s', e)
        raise ValueError(f'Race {event_name!r} on {date} already exists') from e

    cur.execute('SELECT id FROM Races WHERE eventID = ? AND date = ? LIMIT 1', (event_id, date))
    race_id = cur.fetchone()[0]

    with open(csv_filename) as csvfile:
        rows = list(csv.DictReader(csvfile))

    for row in rows:
        try:
            cur.execute(
                'INSERT INTO Athletes (name, age_group, club, hometown) VALUES (?,?,?,?)',
                (row['name'], row['age_group'], row['club'], row['hometown'])
            )
        except Exception as e:
            logger.error('Error inserting athlete: %s', e)
    conn.commit()

    for row in rows:
        cur.execute(
            'SELECT id FROM Athletes WHERE (name = ? AND hometown = ?) LIMIT 1',
            (row['name'], row['hometown'])
        )
        athlete_id = cur.fetchone()[0]
        try:
            cur.execute(
                'INSERT INTO RaceTimes (athleteID, raceID, str_time, sec_time, pace) VALUES (?,?,?,?,?)',
                (athlete_id, race_id, scrub_timestr(row['time']), secs(row['time']), row['pace'])
            )
        except Exception as e:
            logger.error('Error adding to RaceTimes: %s', e)
            raise
    conn.commit()


def get_race_info(cur, race_id: int) -> tuple[str, str]:
    """Return (eventname, date) for a given race id."""
    cur.execute(
        'SELECT eventname, date FROM Races JOIN Events WHERE Races.id = ? AND Events.id = Races.eventID',
        (race_id,)
    )
    result = cur.fetchall()
    return result[0][0], result[0][1]


def get_races_for_athlete(cur, name: str) -> tuple[list, list, list, str]:
    """Return intro/thead/tbody/summary of race results for a named athlete."""
    cur.execute(
        'SELECT eventname, date, str_time, pace FROM RaceTimes '
        'JOIN Athletes ON RaceTimes.athleteID = Athletes.id '
        'JOIN Races ON RaceTimes.raceID = Races.id '
        'JOIN Events ON Events.id = Races.eventID '
        'WHERE Athletes.name = ? ORDER BY date DESC',
        (name,)
    )
    intro = [name]
    thead = ['Event', 'Date', 'Time', 'Pace']
    tbody = [row for row in cur]
    summary = f'Total of {len(tbody)} Races'
    return intro, thead, tbody, summary


def compare_races(cur, event_name1: str, date1: str, event_name2: str, date2: str,
                  race_time1: str, race_time2: str, max_percent: float) -> tuple[tuple, list, list, str]:
    """Compare athletes who ran two races with similar times."""
    intro = (
        'People who ran both races:',
        f'[{event_name1} on {date1}] and [{event_name2} on {date2}]',
        f'with times between {race_time1} and {race_time2}',
        f'where difference in time is less than {int(max_percent * 100)}%',
    )
    cur.execute('DROP TABLE IF EXISTS race1')
    cur.execute('DROP TABLE IF EXISTS race2')
    cur.execute('DROP TABLE IF EXISTS race_compare')

    cur.execute(
        'CREATE TABLE race1 AS SELECT athleteID, str_time, sec_time, pace FROM RaceTimes '
        'JOIN Races ON Races.id = RaceTimes.raceID '
        'JOIN Events ON Races.eventID = Events.id AND Events.eventname = ? AND Races.date = ?',
        (event_name1, date1)
    )
    rows = cur.execute('SELECT * FROM race1').fetchall()
    logger.debug('race1 rows: %d', len(rows))

    cur.execute(
        'CREATE TABLE race2 AS SELECT athleteID, str_time, sec_time, pace FROM RaceTimes '
        'JOIN Races ON Races.id = RaceTimes.raceID '
        'JOIN Events ON Races.eventID = Events.id AND Events.eventname = ? AND Races.date = ?',
        (event_name2, date2)
    )
    rows = cur.execute('SELECT * FROM race2').fetchall()
    logger.debug('race2 rows: %d', len(rows))

    cur.execute(
        'CREATE TABLE race_compare AS '
        'SELECT race2.athleteID, Athletes.name, '
        'race1.str_time AS race1_time_str, race1.sec_time AS race1_time_sec, '
        'race2.str_time AS race2_time_str, race2.sec_time AS race2_time_sec, '
        'race2.sec_time - race1.sec_time AS diff, '
        '(race2.sec_time-race1.sec_time)*2/CAST((race2.sec_time+race1.sec_time)AS REAL) AS percent_diff '
        'FROM race2 '
        'JOIN race1 ON race2.athleteID = race1.athleteID '
        'JOIN Athletes ON race2.AthleteID = Athletes.id '
        'WHERE race1.str_time BETWEEN ? AND ? '
        'AND race2.str_time BETWEEN ? AND ?',
        (race_time1, race_time2, race_time1, race_time2)
    )

    rows = cur.execute('SELECT * FROM race_compare WHERE abs(percent_diff) < ?', (max_percent,)).fetchall()
    diffs = []
    tbody = []
    for row in rows:
        percent = f'{row[7] * 100:.2f}'
        name_tag = Markup(f'<strong><a href=/user/{urlquote(row[1])}>{escape(row[1])}</a></strong>')
        tbody.append([row[0], name_tag, row[2], row[3], row[4], row[5], row[6], percent])
        diffs.append(row[6])

    mean = np.mean(diffs)
    std = np.std(diffs)
    summary = f'mean = {mean:.2f}, standard deviation = {std:.2f} [seconds]'
    thead = ['ID', 'Name', 'Time 1', 'Secs', 'Time 2', 'Secs', 'Diff', '%']
    return intro, thead, tbody, summary


def get_races(cur) -> list[str]:
    """Return a list of race strings in the format 'Event : Date <id>'."""
    races = []
    cur.execute(
        'SELECT Events.eventname, Events.dist, Races.date, Races.id FROM Events '
        'JOIN Races WHERE Races.eventID = Events.id ORDER BY Events.eventname, Races.date'
    )
    for row in cur:
        races.append(f'{row[0]} : {row[2]} <{row[3]}>')
        logger.debug('get_races row=%s', row)
    return races


def load_database(db_name: str) -> tuple:
    """Open a SQLite database and return (conn, cur)."""
    conn = sqlite3.connect(db_name)
    cur = conn.cursor()
    return conn, cur


def get_log_sum(vals) -> list:
    """Format a row of aggregated log sums into display strings."""
    dist = f'{vals[0]:.1f}'
    intensity = f'{vals[1]:.1f}'
    td_str = str(timedelta(seconds=vals[2]))
    recovery_str = str(timedelta(seconds=vals[3]))
    easy_str = str(timedelta(seconds=vals[4]))
    threshold_str = str(timedelta(seconds=vals[5]))
    interval_str = str(timedelta(seconds=vals[6]))
    repetition_str = str(timedelta(seconds=vals[7]))
    return [dist, intensity, td_str, recovery_str, easy_str, threshold_str, interval_str, repetition_str]


def get_log_sums_byisodate(cur, iso_str: str) -> list:
    """Return aggregated log sums for an ISO week pattern (e.g. '2016-01-%')."""
    cur.execute(
        'SELECT sum(dist), sum(jd_int), sum(time_secs), sum(recovery_secs), '
        'sum(easy_secs), sum(threshold_secs), sum(interval_secs), sum(repetition_secs) '
        'FROM Log WHERE isodate LIKE (?)',
        (iso_str,)
    )
    return get_log_sum(cur.fetchone())


def get_log_sums_bygregdate(cur, date_str: str) -> list:
    """Return aggregated log sums for a Gregorian month pattern (e.g. '2016-01%')."""
    cur.execute(
        'SELECT sum(dist), sum(jd_int), sum(time_secs), sum(recovery_secs), '
        'sum(easy_secs), sum(threshold_secs), sum(interval_secs), sum(repetition_secs) '
        'FROM Log WHERE date LIKE (?)',
        (date_str,)
    )
    return get_log_sum(cur.fetchone())


def get_log_sums_for_all_weeks(conn, cur) -> list:
    """Return one summary row per ISO week in the log."""
    conn.create_function("NODAY", 1, noday)
    conn.create_function("NODAYMATCH", 1, nodaymatch)
    cur.execute(
        'SELECT noday(isodate), nodaymatch(isodate) FROM Log '
        'GROUP BY noday(isodate) ORDER by nodaymatch(isodate) DESC'
    )
    tbody = []
    for noday_str, nodaymatch_str in cur.fetchall():
        tbody.append([noday_str] + get_log_sums_byisodate(cur, nodaymatch_str))
    return tbody


def get_log_sums_for_all_months(conn, cur) -> list:
    """Return one summary row per calendar month in the log."""
    conn.create_function("NODAY", 1, noday)
    conn.create_function("NODAYMATCH", 1, nodaymatch)
    cur.execute(
        'SELECT noday(date), nodaymatch(date) FROM Log '
        'GROUP BY noday(date) ORDER by noday(date) DESC'
    )
    tbody = []
    for noday_str, nodaymatch_str in cur.fetchall():
        tbody.append([noday_str] + get_log_sums_bygregdate(cur, nodaymatch_str))
    return tbody


def get_log_sums_over_months(cur, month_ago_nearest: int, month_ago_furthest: int) -> list:
    """Return averaged log sums over a range of calendar months."""
    months = month_ago_furthest - month_ago_nearest + 1

    cur.execute('SELECT noday(date) FROM Log GROUP BY noday(date) ORDER by noday(date) DESC')
    curlist = cur.fetchall()

    start_date = curlist[month_ago_furthest][0]
    end_date = curlist[month_ago_nearest][0]
    start_date_match = start_date + '-01'
    end_date_match = end_date + '-31'

    cur.execute(
        'SELECT sum(dist), sum(jd_int), sum(time_secs), sum(recovery_secs), sum(easy_secs), '
        'sum(threshold_secs), sum(interval_secs), sum(repetition_secs) '
        'FROM Log WHERE date BETWEEN ? and ?',
        (start_date_match, end_date_match)
    )
    vals = cur.fetchone()

    total_secs = vals[2]
    recovery_secs = vals[3]
    easy_secs = vals[4]
    p8020 = f'{(100.0 * (recovery_secs + easy_secs) / total_secs) if total_secs else 0.0:.1f}'

    dist = f'{vals[0] / months:.1f}'
    intensity = f'{vals[1] / months:.1f}'
    td_str = str(timedelta(seconds=int(vals[2] / months)))
    recovery_str = str(timedelta(seconds=int(vals[3] / months)))
    easy_str = str(timedelta(seconds=int(vals[4] / months)))
    threshold_str = str(timedelta(seconds=int(vals[5] / months)))
    interval_str = str(timedelta(seconds=int(vals[6] / months)))
    repetition_str = str(timedelta(seconds=int(vals[7] / months)))

    zones = zones_str(recovery_str, easy_str, threshold_str, interval_str, repetition_str)
    return [dist, td_str, zones, intensity, p8020]


def get_dist_by_type(cur, dist_type: str, date_type: str,
                     start_date_match: str, end_date_match: str) -> float:
    """Return total distance for a workout type over a date range."""
    if date_type == 'Isodate':
        cur.execute(
            'SELECT sum(dist) FROM Log JOIN wo_type ON Log.wo_type = wo_type.id '
            'WHERE isodate between ? and ? and type == ?',
            (start_date_match, end_date_match, dist_type)
        )
    elif date_type == 'Date':
        cur.execute(
            'SELECT sum(dist) FROM Log JOIN wo_type ON Log.wo_type = wo_type.id '
            'WHERE date between ? and ? and type == ?',
            (start_date_match, end_date_match, dist_type)
        )
    else:
        raise ValueError(f'Unknown date_type {date_type!r}: expected "Isodate" or "Date"')

    dist = cur.fetchone()[0]
    return dist if dist is not None else 0.0


def get_log_sums_over_weeks(cur, week_ago_nearest: int, week_ago_furthest: int) -> list:
    """Return averaged log sums over a range of ISO weeks."""
    weeks = week_ago_furthest - week_ago_nearest + 1

    cur.execute('SELECT noday(isodate) FROM Log GROUP BY noday(isodate) ORDER by noday(isodate) DESC')
    curlist = cur.fetchall()

    start_date = curlist[week_ago_furthest][0]
    end_date = curlist[week_ago_nearest][0]
    start_date_match = start_date + '-1'
    end_date_match = end_date + '-7'

    run_dist = f'{get_dist_by_type(cur, "Run", "Isodate", start_date_match, end_date_match) / weeks:.1f}'
    ride_dist = f'{get_dist_by_type(cur, "Ride", "Isodate", start_date_match, end_date_match) / weeks:.1f}'

    cur.execute(
        'SELECT sum(dist), sum(jd_int), sum(time_secs), sum(recovery_secs), sum(easy_secs), '
        'sum(threshold_secs), sum(interval_secs), sum(repetition_secs) '
        'FROM Log WHERE isodate BETWEEN ? and ?',
        (start_date_match, end_date_match)
    )
    vals = cur.fetchone()

    total_secs = vals[2]
    recovery_secs = vals[3]
    easy_secs = vals[4]
    p8020 = f'{(100.0 * (recovery_secs + easy_secs) / total_secs) if total_secs else 0.0:.1f}'

    dist = Markup(
        f'<strong>Run: </strong>{run_dist}'
        f'<br /><strong>Ride: </strong>{ride_dist}'
        f'<br /><strong>Total: </strong>{vals[0] / weeks:.1f}'
    )
    intensity = f'{vals[1] / weeks:.1f}'
    td_str = str(timedelta(seconds=int(vals[2] / weeks)))
    recovery_str = str(timedelta(seconds=int(vals[3] / weeks)))
    easy_str = str(timedelta(seconds=int(vals[4] / weeks)))
    threshold_str = str(timedelta(seconds=int(vals[5] / weeks)))
    interval_str = str(timedelta(seconds=int(vals[6] / weeks)))
    repetition_str = str(timedelta(seconds=int(vals[7] / weeks)))

    zones = zones_str(recovery_str, easy_str, threshold_str, interval_str, repetition_str)
    return [dist, td_str, zones, intensity, p8020]


def _register_date_functions(conn) -> None:
    """Register NODAY and NODAYMATCH SQLite user-defined functions."""
    conn.create_function("NODAY", 1, noday)
    conn.create_function("NODAYMATCH", 1, nodaymatch)


def _build_stats_table(conn, cur, title: str, sums_fn, rows) -> tuple[list, list, list, str]:
    """Build a stats table by calling sums_fn for each (label, near, far) row."""
    _register_date_functions(conn)
    thead = ['Description', 'Dist.', 'Time', 'Time in Zones', 'Int.', '%80/20']
    tbody = [[label] + sums_fn(cur, near, far) for label, near, far in rows]
    return [title], thead, tbody, ''


def get_log_52weeks(conn, cur) -> tuple[list, list, list, str]:
    """Return weekly stats for the last 52 weeks."""
    return _build_stats_table(conn, cur, 'Weekly Stats', get_log_sums_over_weeks,
                              [(f'{w} weeks ago', w, w) for w in range(1, 53)])


def get_log_week_stats(conn, cur) -> tuple[list, list, list, str]:
    """Return weekly stats with rolling averages."""
    return _build_stats_table(conn, cur, 'Weekly Stats', get_log_sums_over_weeks, [
        ('This Week',          0,  0),
        ('Last Week',          1,  1),
        ('Two Weeks Ago',      2,  2),
        ('Three Weeks Ago',    3,  3),
        ('Four Weeks Ago',     4,  4),
        ('Ave: Last 2 weeks',  1,  2),
        ('Ave: Last 4 weeks',  1,  4),
        ('Ave: Last 8 weeks',  1,  8),
        ('Ave: Last 16 weeks', 1, 16),
        ('Ave: Last 32 weeks', 1, 32),
        ('Ave: Last 52 weeks', 1, 52),
    ])


def get_log_12months(conn, cur) -> tuple[list, list, list, str]:
    """Return monthly stats for the last 12 months."""
    return _build_stats_table(conn, cur, 'Monthly Stats', get_log_sums_over_months,
                              [(f'{m} months ago', m, m) for m in range(1, 13)])


def get_log_month_stats(conn, cur) -> tuple[list, list, list, str]:
    """Return monthly stats with rolling averages."""
    return _build_stats_table(conn, cur, 'Monthly Stats', get_log_sums_over_months, [
        ('This Month',          0,  0),
        ('Last Month',          1,  1),
        ('Two Months Ago',      2,  2),
        ('Three Months Ago',    3,  3),
        ('Four Months Ago',     4,  4),
        ('Ave: Last 2 months',  1,  2),
        ('Ave: Last 3 months',  1,  3),
        ('Ave: Last 4 months',  1,  4),
        ('Ave: Last 5 months',  1,  5),
        ('Ave: Last 6 months',  1,  6),
        ('Ave: Last 8 months',  1,  8),
        ('Ave: Last 12 months', 1, 12),
    ])


def get_athletes(conn, cur, letter: str | None = None) -> tuple[list, list, list, str]:
    """Return intro/thead/tbody/summary for athletes, optionally filtered by first letter."""
    if letter is not None:
        cur.execute(
            'SELECT id, name, age_group, club, hometown FROM Athletes '
            'WHERE name LIKE ? ORDER by name LIMIT 10 OFFSET 0',
            (f'{letter}%',)
        )
    else:
        cur.execute('SELECT id, name, age_group, club, hometown FROM Athletes ORDER by name')

    tbody = []
    for athlete_id, name, age_group, club, hometown in cur:
        id_tag = Markup(f'<strong><a href=/user/{urlquote(name)}>{athlete_id}</a></strong>')
        tbody.append([id_tag, name, age_group, club, hometown])

    intro = ''
    thead = ['id', 'name', 'age_group', 'club', 'hometown']
    summary = f'Total of {len(tbody)} Entries'
    return intro, thead, tbody, summary


def noday(date_str: str) -> str:
    """Strip the day from a date/isodate string, returning 'YYYY-MM'."""
    year, mw, _ = parse_datestr(date_str)
    return f'{year}-{mw}'


def nodaymatch(date_str: str) -> str:
    """Return a LIKE pattern matching all days in a month/week: 'YYYY-MM-%'."""
    year, mw, _ = parse_datestr(date_str)
    return f'{year}-{mw}-%'


def norm_race_elev(time_str: str, elev: float, dist: float) -> str:
    """Return elevation-adjusted race time string."""
    time_sec = secs(time_str)
    ave_grade = elev / (dist * 1000)
    percent_increase = 1.5 * ave_grade
    adjusted_time = (1 - percent_increase) * time_sec
    return str_time(adjusted_time)


def norm_race_dist(time_str: str, orig_dist: float, norm_dist: float) -> str:
    """Return distance-normalised race time string."""
    time_sec = secs(time_str)
    time_sec = time_sec * (norm_dist / orig_dist) ** 1.06
    return str_time(time_sec)


def main_sql() -> None:
    from config import DATABASE
    conn, cur = load_database(DATABASE)
    csv_filename = "/Users/sroberts/GitHub/Running/health_221226.csv"
    load_csv_health(conn, cur, csv_filename)
    conn.commit()
    cur.close()


if __name__ == "__main__":
    main_sql()
