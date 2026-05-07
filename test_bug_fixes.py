#!/usr/bin/env python3
"""Tests for the 6 high-priority bug fixes in sql.py and sql_maintenance.py."""

import unittest
import sqlite3
import datetime
import sys
import os
from unittest.mock import patch

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import sql
import sql_maintenance
import races
from races import app, _validate_workout, _validate_health, _validate_shoe


def make_db():
    """Return (conn, cur) for an in-memory SQLite database with the full schema."""
    conn = sqlite3.connect(':memory:')
    cur = conn.cursor()

    cur.execute('CREATE TABLE wo_type (id INTEGER PRIMARY KEY, type TEXT)')
    cur.execute('CREATE TABLE Shoes (id INTEGER PRIMARY KEY, shortName TEXT, longName TEXT, retired INTEGER)')
    cur.execute('''CREATE TABLE Log (
        id INTEGER PRIMARY KEY, date TEXT, location TEXT, wo_type INTEGER,
        objective TEXT, notes TEXT, dist REAL, time TEXT, time_secs REAL,
        pace TEXT, pace_secs REAL, recovery TEXT, recovery_secs REAL,
        easy TEXT, easy_secs REAL, threshold TEXT, threshold_secs REAL,
        interval TEXT, interval_secs REAL, repetition TEXT, repetition_secs REAL,
        p8020 REAL, jd_int REAL, shoeID INTEGER, isodate TEXT
    )''')
    cur.execute('''CREATE TABLE Events (
        id INTEGER PRIMARY KEY, eventname TEXT, dist REAL,
        min_elev INTEGER, max_elev INTEGER, gain_elev INTEGER, UNIQUE (eventname)
    )''')
    cur.execute('CREATE TABLE Races (id INTEGER PRIMARY KEY, eventID INTEGER, date TEXT, UNIQUE (eventID, date))')
    cur.execute('''CREATE TABLE Athletes (
        id INTEGER PRIMARY KEY, name TEXT, age_group TEXT, club TEXT, hometown TEXT,
        UNIQUE (name, hometown)
    )''')
    cur.execute('CREATE TABLE RaceTimes (id INTEGER PRIMARY KEY, athleteID INTEGER, raceID INTEGER, str_time TEXT, sec_time INTEGER, pace TEXT)')
    cur.execute('''CREATE TABLE Health (
        id INTEGER PRIMARY KEY, date TEXT, weight REAL, smallWaist REAL, bbWaist REAL,
        hip REAL, chest REAL, notes TEXT, HR REAL, UNIQUE (date)
    )''')

    cur.execute("INSERT INTO wo_type VALUES (1, 'Run')")
    cur.execute("INSERT INTO wo_type VALUES (2, 'Ride')")
    cur.execute("INSERT INTO Shoes VALUES (1, 'TestShoe', 'Test Shoe Long', 0)")
    conn.commit()
    return conn, cur


def add_wo(conn, cur, date='2024-01-15', location='Loc', wo_type_id=1,
           objective='obj', notes='notes', distance='10.0',
           recovery='0:00:00', easy='1:00:00', threshold='0:00:00',
           interval='0:00:00', repetition='0:00:00', shoe_id=1):
    sql.add_workout(conn, cur, date, location, wo_type_id, objective, notes,
                    distance, recovery, easy, threshold, interval, repetition, shoe_id)


# ---------------------------------------------------------------------------
# BUG-1: Division by zero in p8020 calculation
# ---------------------------------------------------------------------------
class TestBug1ZeroDivision(unittest.TestCase):

    def test_all_zones_zero_returns_zero_p8020(self):
        """Zero total time must return p8020=0.0, not raise ZeroDivisionError."""
        wd = sql.get_workout_calculated_data(
            '2024-01-15', '0.0',
            '0:00:00', '0:00:00', '0:00:00', '0:00:00', '0:00:00'
        )
        self.assertEqual(wd.p8020, 0.0)

    def test_easy_only_gives_100_percent(self):
        wd = sql.get_workout_calculated_data(
            '2024-01-15', '10.0',
            '0:00:00', '1:00:00', '0:00:00', '0:00:00', '0:00:00'
        )
        self.assertAlmostEqual(wd.p8020, 100.0)

    def test_mixed_zones_correct_ratio(self):
        # 60 min easy+recovery out of 90 min total = 66.67 %
        wd = sql.get_workout_calculated_data(
            '2024-01-15', '10.0',
            '0:30:00', '0:30:00', '0:30:00', '0:00:00', '0:00:00'
        )
        self.assertAlmostEqual(wd.p8020, 66.67, places=1)

    def test_get_log_sums_over_weeks_zero_time(self):
        """get_log_sums_over_weeks must not crash when all time zones are zero."""
        conn, cur = make_db()
        conn.create_function("NODAY", 1, sql.noday)
        add_wo(conn, cur, easy='0:00:00')
        result = sql.get_log_sums_over_weeks(cur, 0, 0)
        self.assertEqual(result[-1], '0.0')
        conn.close()

    def test_get_log_sums_over_months_zero_time(self):
        """get_log_sums_over_months must not crash when all time zones are zero."""
        conn, cur = make_db()
        conn.create_function("NODAY", 1, sql.noday)
        add_wo(conn, cur, easy='0:00:00')
        result = sql.get_log_sums_over_months(cur, 0, 0)
        self.assertEqual(result[-1], '0.0')
        conn.close()


# ---------------------------------------------------------------------------
# BUG-2: update_workout_calculated_data passes wrong args to change_workout
# ---------------------------------------------------------------------------
class TestBug2MissingWoTypeId(unittest.TestCase):

    def test_wo_type_preserved_after_update(self):
        """update_workout_calculated_data must not corrupt wo_type field."""
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15', wo_type_id=1)
        add_wo(conn, cur, date='2024-01-16', wo_type_id=2)

        sql_maintenance.update_workout_calculated_data(conn, cur)

        cur.execute('SELECT wo_type FROM Log ORDER BY date')
        types = [row[0] for row in cur.fetchall()]
        self.assertEqual(types, [1, 2])
        conn.close()

    def test_other_fields_preserved_after_update(self):
        """update_workout_calculated_data must not change location or objective."""
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15', location='TrackA', objective='Intervals')

        sql_maintenance.update_workout_calculated_data(conn, cur)

        cur.execute('SELECT location, objective FROM Log')
        row = cur.fetchone()
        self.assertEqual(row[0], 'TrackA')
        self.assertEqual(row[1], 'Intervals')
        conn.close()


# ---------------------------------------------------------------------------
# BUG-3: nodaymatch never registered as SQLite function
# ---------------------------------------------------------------------------
class TestBug3NodeaymatchNotRegistered(unittest.TestCase):
    """
    nodaymatch() is used as a SQLite user-defined function in
    get_log_sums_for_all_weeks and get_log_sums_for_all_months, but was never
    registered via conn.create_function(). The four public stat functions also
    now defensively register it.
    """

    def test_nodaymatch_python_function(self):
        self.assertEqual(sql.nodaymatch('2024-03-15'), '2024-03-%')
        self.assertEqual(sql.nodaymatch('2024-01-01'), '2024-01-%')

    def _make_db_with_many_workouts(self):
        """Seed 60 weekly workouts so all stat range look-ups succeed."""
        from datetime import date, timedelta
        conn, cur = make_db()
        base = date(2023, 1, 2)   # Monday
        for i in range(60):
            d = (base + timedelta(weeks=i)).strftime('%Y-%m-%d')
            add_wo(conn, cur, date=d)
        return conn, cur

    # --- get_log_sums_for_all_weeks / months: the functions that actually use
    #     nodaymatch() inside a SQL query ---

    def test_get_log_sums_for_all_weeks_no_crash(self):
        """get_log_sums_for_all_weeks must not raise OperationalError on nodaymatch."""
        conn, cur = self._make_db_with_many_workouts()
        try:
            sql.get_log_sums_for_all_weeks(conn, cur)
        except Exception as e:
            self.fail(f'get_log_sums_for_all_weeks raised: {e}')
        conn.close()

    def test_get_log_sums_for_all_months_no_crash(self):
        """get_log_sums_for_all_months must not raise OperationalError on nodaymatch."""
        conn, cur = self._make_db_with_many_workouts()
        try:
            sql.get_log_sums_for_all_months(conn, cur)
        except Exception as e:
            self.fail(f'get_log_sums_for_all_months raised: {e}')
        conn.close()

    # --- Public stat functions also register nodaymatch defensively ---

    def test_get_log_52weeks_no_crash(self):
        conn, cur = self._make_db_with_many_workouts()
        try:
            sql.get_log_52weeks(conn, cur)
        except Exception as e:
            self.fail(f'get_log_52weeks raised: {e}')
        conn.close()

    def test_get_log_week_stats_no_crash(self):
        conn, cur = self._make_db_with_many_workouts()
        try:
            sql.get_log_week_stats(conn, cur)
        except Exception as e:
            self.fail(f'get_log_week_stats raised: {e}')
        conn.close()

    def test_get_log_12months_no_crash(self):
        conn, cur = self._make_db_with_many_workouts()
        try:
            sql.get_log_12months(conn, cur)
        except Exception as e:
            self.fail(f'get_log_12months raised: {e}')
        conn.close()

    def test_get_log_month_stats_no_crash(self):
        conn, cur = self._make_db_with_many_workouts()
        try:
            sql.get_log_month_stats(conn, cur)
        except Exception as e:
            self.fail(f'get_log_month_stats raised: {e}')
        conn.close()


# ---------------------------------------------------------------------------
# BUG-4: compare_races iterates race1 rows when displaying race2
# ---------------------------------------------------------------------------
class TestBug4CompareRacesStaleRows(unittest.TestCase):

    def _populate(self, conn, cur):
        cur.execute("INSERT INTO Events VALUES (1, 'Race A', 10.0, 0, 0, 0)")
        cur.execute("INSERT INTO Events VALUES (2, 'Race B', 10.0, 0, 0, 0)")
        cur.execute("INSERT INTO Races VALUES (1, 1, '2024-01-01')")
        cur.execute("INSERT INTO Races VALUES (2, 2, '2024-06-01')")
        # Alice ran both; Bob only Race A
        cur.execute("INSERT INTO Athletes VALUES (1, 'Alice', 'F30', 'TC', 'Victoria')")
        cur.execute("INSERT INTO Athletes VALUES (2, 'Bob', 'M30', 'TC', 'Victoria')")
        cur.execute("INSERT INTO RaceTimes VALUES (1, 1, 1, '0:40:00', 2400, '4:00')")
        cur.execute("INSERT INTO RaceTimes VALUES (2, 2, 1, '0:41:00', 2460, '4:06')")
        cur.execute("INSERT INTO RaceTimes VALUES (3, 1, 2, '0:39:00', 2340, '3:54')")
        conn.commit()

    def test_only_athlete_who_ran_both_appears(self):
        conn, cur = make_db()
        self._populate(conn, cur)
        result = sql.compare_races(cur, 'Race A', '2024-01-01', 'Race B', '2024-06-01',
                                   '0:35:00', '0:45:00', 0.5)
        tbody = result[2]
        self.assertEqual(len(tbody), 1, "Only Alice ran both races")
        self.assertEqual(tbody[0][0], 1, "Row should be for Alice (athleteID=1)")
        conn.close()

    def test_race2_times_are_used_not_race1(self):
        """race2 column in result should contain Race B time, not Race A time."""
        conn, cur = make_db()
        self._populate(conn, cur)
        result = sql.compare_races(cur, 'Race A', '2024-01-01', 'Race B', '2024-06-01',
                                   '0:35:00', '0:45:00', 0.5)
        tbody = result[2]
        # tbody[0] = [athleteID, name_tag, race1_time_str, race1_sec, race2_time_str, race2_sec, diff, pct]
        race2_time_str = tbody[0][4]
        self.assertEqual(race2_time_str, '0:39:00', "race2 time should be Alice's Race B time")
        conn.close()


# ---------------------------------------------------------------------------
# BUG-5: get_dist_by_type "Date" branch queries isodate instead of date
# ---------------------------------------------------------------------------
class TestBug5GetDistByTypeDateBranch(unittest.TestCase):

    def test_date_branch_filters_on_gregorian_date(self):
        """'Date' datetype must use the date column, not isodate."""
        conn, cur = make_db()
        # Two workouts in different Gregorian months
        add_wo(conn, cur, date='2024-01-15', distance='10.0')
        add_wo(conn, cur, date='2024-06-15', distance='20.0')

        dist = sql.get_dist_by_type(cur, 'Run', 'Date', '2024-01-01', '2024-01-31')
        self.assertAlmostEqual(dist, 10.0)

        dist = sql.get_dist_by_type(cur, 'Run', 'Date', '2024-06-01', '2024-06-30')
        self.assertAlmostEqual(dist, 20.0)
        conn.close()

    def test_date_branch_excludes_out_of_range(self):
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15', distance='10.0')

        dist = sql.get_dist_by_type(cur, 'Run', 'Date', '2024-06-01', '2024-06-30')
        self.assertAlmostEqual(dist, 0.0)
        conn.close()

    def test_isodate_branch_unchanged(self):
        """Isodate branch must still filter on the isodate column."""
        conn, cur = make_db()
        # Jan 15, 2024 is ISO week 2024-03
        add_wo(conn, cur, date='2024-01-15', distance='10.0')

        dist = sql.get_dist_by_type(cur, 'Run', 'Isodate', '2024-03-1', '2024-03-7')
        self.assertAlmostEqual(dist, 10.0)
        conn.close()


# ---------------------------------------------------------------------------
# BUG-6: read_ss passes objective where wo_type_id is expected
# ---------------------------------------------------------------------------
class TestBug6ReadSsMissingWoTypeId(unittest.TestCase):

    def _make_sheet(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        # read_ss starts at row 5; col layout:
        # 1=date 2=location 3=objective 4=notes 5=distance 6=time 15=shoes
        ws.cell(row=5, column=1).value = '2024-01-15'
        ws.cell(row=5, column=2).value = 'Test Location'
        ws.cell(row=5, column=3).value = 'Test Objective'
        ws.cell(row=5, column=4).value = 'Test Notes'
        ws.cell(row=5, column=5).value = 10.0
        ws.cell(row=5, column=6).value = '1:00:00'
        ws.cell(row=5, column=15).value = 'TestShoe'
        return ws

    def test_workout_inserted_with_valid_wo_type(self):
        conn, cur = make_db()
        sql.read_ss(self._make_sheet(), conn, cur)
        cur.execute('SELECT wo_type FROM Log')
        row = cur.fetchone()
        self.assertIsNotNone(row, "A workout should have been inserted")
        self.assertIsInstance(row[0], int, "wo_type should be an integer ID")
        conn.close()

    def test_other_fields_are_correct(self):
        conn, cur = make_db()
        sql.read_ss(self._make_sheet(), conn, cur)
        cur.execute('SELECT date, location, objective FROM Log')
        row = cur.fetchone()
        self.assertEqual(row[0], '2024-01-15')
        self.assertEqual(row[1], 'Test Location')
        self.assertEqual(row[2], 'Test Objective')
        conn.close()


# ---------------------------------------------------------------------------
# SEC-1: XSS via Markup() on unsanitized DB values
# ---------------------------------------------------------------------------
class TestSec1XssPrevention(unittest.TestCase):
    """DB values containing HTML must be escaped before being rendered."""

    XSS_PAYLOAD = '<script>alert(1)</script>'

    def test_get_workouts_description_escapes_location(self):
        conn, cur = make_db()
        add_wo(conn, cur, location=self.XSS_PAYLOAD)
        [_, _, tbody, _] = sql.get_workouts(cur)
        description = str(tbody[0][2])
        self.assertNotIn('<script>', description)
        self.assertIn('&lt;script&gt;', description)
        conn.close()

    def test_get_workouts_description_escapes_objective(self):
        conn, cur = make_db()
        add_wo(conn, cur, objective=self.XSS_PAYLOAD)
        [_, _, tbody, _] = sql.get_workouts(cur)
        description = str(tbody[0][2])
        self.assertNotIn('<script>', description)
        self.assertIn('&lt;script&gt;', description)
        conn.close()

    def test_get_workouts_description_escapes_notes(self):
        conn, cur = make_db()
        add_wo(conn, cur, notes=self.XSS_PAYLOAD)
        [_, _, tbody, _] = sql.get_workouts(cur)
        description = str(tbody[0][2])
        self.assertNotIn('<script>', description)
        self.assertIn('&lt;script&gt;', description)
        conn.close()

    def test_zones_str_escapes_values(self):
        """zones_str must escape time strings that contain HTML characters."""
        result = str(sql.zones_str(
            '0:00:00', '<b>bad</b>', '0:00:00', '0:00:00', '0:00:00'
        ))
        self.assertNotIn('<b>', result)
        self.assertIn('&lt;b&gt;', result)

    def test_get_health_list_id_tag_is_relative_url(self):
        conn, cur = make_db()
        sql.add_health(conn, cur, '2024-01-15', 70.0, 80.0, 85.0, 90.0, 95.0, 'note', 60)
        [_, _, tbody, _] = sql.get_health_list(cur)
        id_tag = str(tbody[0][0])
        self.assertNotIn('localhost', id_tag)
        self.assertIn('/ChangeHealth/', id_tag)
        conn.close()


# ---------------------------------------------------------------------------
# SEC-2: No hardcoded localhost:8080 in any returned HTML
# ---------------------------------------------------------------------------
class TestSec2NoHardcodedUrls(unittest.TestCase):
    """sql.py must never return localhost:8080 URLs — all links are relative."""

    def test_get_workouts_id_tag_is_relative(self):
        conn, cur = make_db()
        add_wo(conn, cur)
        [_, _, tbody, _] = sql.get_workouts(cur)
        id_tag = str(tbody[0][0])
        self.assertNotIn('localhost', id_tag)
        self.assertIn('/ChangeWorkout/', id_tag)
        conn.close()

    def test_get_athletes_id_tag_is_relative(self):
        conn, cur = make_db()
        cur.execute("INSERT INTO Athletes VALUES (1, 'Alice Smith', 'F30', 'TC', 'Victoria')")
        conn.commit()
        [_, _, tbody, _] = sql.get_athletes(conn, cur)
        id_tag = str(tbody[0][0])
        self.assertNotIn('localhost', id_tag)
        self.assertIn('/user/', id_tag)
        conn.close()

    def test_get_athletes_name_with_spaces_is_url_encoded(self):
        conn, cur = make_db()
        cur.execute("INSERT INTO Athletes VALUES (1, 'Alice Smith', 'F30', 'TC', 'Victoria')")
        conn.commit()
        [_, _, tbody, _] = sql.get_athletes(conn, cur)
        id_tag = str(tbody[0][0])
        self.assertIn('Alice%20Smith', id_tag)
        conn.close()

    def test_compare_races_name_tag_is_relative(self):
        conn, cur = make_db()
        cur.execute("INSERT INTO Events VALUES (1, 'Race A', 10.0, 0, 0, 0)")
        cur.execute("INSERT INTO Events VALUES (2, 'Race B', 10.0, 0, 0, 0)")
        cur.execute("INSERT INTO Races VALUES (1, 1, '2024-01-01')")
        cur.execute("INSERT INTO Races VALUES (2, 2, '2024-06-01')")
        cur.execute("INSERT INTO Athletes VALUES (1, 'Bob Jones', 'M30', 'TC', 'Victoria')")
        cur.execute("INSERT INTO RaceTimes VALUES (1, 1, 1, '0:40:00', 2400, '4:00')")
        cur.execute("INSERT INTO RaceTimes VALUES (2, 1, 2, '0:39:00', 2340, '3:54')")
        conn.commit()
        result = sql.compare_races(cur, 'Race A', '2024-01-01', 'Race B', '2024-06-01',
                                   '0:35:00', '0:45:00', 0.5)
        tbody = result[2]
        if tbody:
            name_tag = str(tbody[0][1])
            self.assertNotIn('localhost', name_tag)
            self.assertIn('/user/', name_tag)
        conn.close()

    def test_compare_races_name_escapes_xss(self):
        """Athlete name containing HTML must be escaped in the link text."""
        conn, cur = make_db()
        xss_name = 'Bob<script>alert(1)</script>'
        cur.execute("INSERT INTO Events VALUES (1, 'Race A', 10.0, 0, 0, 0)")
        cur.execute("INSERT INTO Events VALUES (2, 'Race B', 10.0, 0, 0, 0)")
        cur.execute("INSERT INTO Races VALUES (1, 1, '2024-01-01')")
        cur.execute("INSERT INTO Races VALUES (2, 2, '2024-06-01')")
        cur.execute("INSERT INTO Athletes VALUES (1, ?, 'M30', 'TC', 'Victoria')", (xss_name,))
        cur.execute("INSERT INTO RaceTimes VALUES (1, 1, 1, '0:40:00', 2400, '4:00')")
        cur.execute("INSERT INTO RaceTimes VALUES (2, 1, 2, '0:39:00', 2340, '3:54')")
        conn.commit()
        result = sql.compare_races(cur, 'Race A', '2024-01-01', 'Race B', '2024-06-01',
                                   '0:35:00', '0:45:00', 0.5)
        tbody = result[2]
        if tbody:
            name_tag = str(tbody[0][1])
            self.assertNotIn('<script>', name_tag)
        conn.close()


# ---------------------------------------------------------------------------
# CODE-1: write_ss saves workbook on every row iteration
# ---------------------------------------------------------------------------
class TestCode1WriteSsSaveOnce(unittest.TestCase):
    """wb.save() must be called exactly once, after all rows are written.

    write_ss calls _get_workouts_raw internally; we mock it to control the
    data shape and isolate the save-count behaviour.
    """

    # 15 raw columns: date, location, objective, notes, dist, time, pace,
    # recovery, easy, threshold, interval, repetition, p8020, jd_int, shortName
    _FAKE_ROW = ('2024-01-10', 'Loc', 'obj', 'notes', 10.0,
                 '1:00:00', '6:00', '0:00:00', '1:00:00', '0:00:00',
                 '0:00:00', '0:00:00', 100.0, 12.0, 'TestShoe')
    _FAKE_ROWS = [_FAKE_ROW, _FAKE_ROW, _FAKE_ROW]

    def test_save_called_once_for_multiple_rows(self):
        import tempfile, openpyxl
        from unittest.mock import patch

        conn, cur = make_db()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            xlfile = f.name

        save_calls = []
        original_save = openpyxl.workbook.workbook.Workbook.save

        def counting_save(self_wb, filename):
            save_calls.append(filename)
            original_save(self_wb, filename)

        with patch.object(openpyxl.workbook.workbook.Workbook, 'save', counting_save), \
             patch.object(sql, '_get_workouts_raw', return_value=self._FAKE_ROWS):
            sql.write_ss(conn, cur, xlfile)

        self.assertEqual(len(save_calls), 1,
                         f"wb.save() should be called once, was called {len(save_calls)} times")
        conn.close()

    def test_output_file_contains_all_rows(self):
        import tempfile, openpyxl
        from unittest.mock import patch

        conn, cur = make_db()
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            xlfile = f.name

        with patch.object(sql, '_get_workouts_raw', return_value=self._FAKE_ROWS):
            sql.write_ss(conn, cur, xlfile)

        wb = openpyxl.load_workbook(xlfile)
        ws = wb.active
        # Row 4 is the header; data starts at row 5
        data_rows = [ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)]
        non_empty = [v for v in data_rows if v is not None]
        self.assertEqual(len(non_empty), 3)
        conn.close()


# ---------------------------------------------------------------------------
# LOW-1: DB path centralised in config.py
# ---------------------------------------------------------------------------
class TestLow1CentralisedDbPath(unittest.TestCase):
    """All modules must read the DB path from config.DATABASE."""

    def test_config_module_exports_database(self):
        import config
        self.assertTrue(hasattr(config, 'DATABASE'), "config.py must export DATABASE")
        self.assertIsInstance(config.DATABASE, str)
        self.assertTrue(config.DATABASE.endswith('.sqlite'), "DATABASE should point to a .sqlite file")

    def test_env_var_overrides_default(self):
        import importlib, config as cfg_module
        original = cfg_module.DATABASE
        try:
            with __import__('unittest.mock', fromlist=['patch']).patch.dict(
                    __import__('os').environ, {'RUNNING_DB': '/tmp/test.sqlite'}):
                import importlib
                importlib.reload(cfg_module)
                self.assertEqual(cfg_module.DATABASE, '/tmp/test.sqlite')
        finally:
            importlib.reload(cfg_module)

    def test_races_py_uses_config(self):
        import ast, pathlib
        src = pathlib.Path('races.py').read_text()
        tree = ast.parse(src)
        hardcoded = [
            node for node in ast.walk(tree)
            if isinstance(node, ast.Constant)
            and isinstance(node.value, str)
            and 'sqlite' in node.value
        ]
        self.assertEqual(hardcoded, [], f"races.py has hardcoded DB path(s): {[n.value for n in hardcoded]}")

    def test_sql_maintenance_uses_config(self):
        import ast, pathlib
        src = pathlib.Path('sql_maintenance.py').read_text()
        tree = ast.parse(src)
        hardcoded = [
            node for node in ast.walk(tree)
            if isinstance(node, ast.Constant)
            and isinstance(node.value, str)
            and 'sqlite' in node.value
        ]
        self.assertEqual(hardcoded, [], f"sql_maintenance.py has hardcoded DB path(s): {[n.value for n in hardcoded]}")


# ---------------------------------------------------------------------------
# CODE-9: Stale val_date = '2016-MM-DD' placeholder
# ---------------------------------------------------------------------------
class TestCode9ValDate(unittest.TestCase):

    def test_val_date_is_not_stale_placeholder(self):
        """AddHealth and AddWorkout must not seed val_date with a 2016 literal."""
        import ast, pathlib
        src = pathlib.Path('races.py').read_text()
        tree = ast.parse(src)
        stale = [
            node for node in ast.walk(tree)
            if isinstance(node, ast.Constant) and node.value == '2016-MM-DD'
        ]
        self.assertEqual(stale, [], "Found stale '2016-MM-DD' literal in races.py")

    def test_val_date_format_matches_iso(self):
        """The date produced today must match YYYY-MM-DD."""
        import re
        val_date = datetime.date.today().strftime('%Y-%m-%d')
        self.assertRegex(val_date, r'^\d{4}-\d{2}-\d{2}$')


# ---------------------------------------------------------------------------
# CODE-2: secs/str_time shadow their own parameter names
# ---------------------------------------------------------------------------
class TestCode2ParamNameShadowing(unittest.TestCase):

    def test_secs_returns_correct_seconds(self):
        self.assertEqual(sql.secs('0:01:00'), 60.0)
        self.assertEqual(sql.secs('1:00:00'), 3600.0)
        self.assertEqual(sql.secs('0:00:30'), 30.0)

    def test_str_time_returns_correct_string(self):
        self.assertEqual(sql.str_time(60), '0:01:00')
        self.assertEqual(sql.str_time(3600), '1:00:00')
        self.assertEqual(sql.str_time(90), '0:01:30')

    def test_secs_and_str_time_are_inverses(self):
        original = '1:23:45'
        self.assertEqual(sql.str_time(sql.secs(original)), original)


# ---------------------------------------------------------------------------
# LOW-2: weekday hardcoded as 'Monday'
# ---------------------------------------------------------------------------
class TestLow2WeekdayDerived(unittest.TestCase):

    def test_weekday_matches_actual_day(self):
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15')  # Monday
        add_wo(conn, cur, date='2024-01-17')  # Wednesday
        add_wo(conn, cur, date='2024-01-20')  # Saturday
        running_dict = sql.get_running_dict(conn, cur)
        days = [running_dict[i]['weekday'] for i in sorted(running_dict)]
        self.assertEqual(days, ['Monday', 'Wednesday', 'Saturday'])
        conn.close()


# ---------------------------------------------------------------------------
# LOW-3: retired IS 0 → retired = 0
# ---------------------------------------------------------------------------
class TestLow3RetiredComparison(unittest.TestCase):

    def test_no_identity_comparison_on_integer(self):
        import ast, pathlib
        src = pathlib.Path('races.py').read_text()
        # Check raw source for the pattern
        self.assertNotIn('retired is 0', src,
                         "SQL query uses 'IS' for integer comparison instead of '='")


# ---------------------------------------------------------------------------
# LOW-4: parse_datestr raises ValueError on bad input
# ---------------------------------------------------------------------------
class TestLow4ParseDatestrErrorHandling(unittest.TestCase):

    def test_valid_date_parsed_correctly(self):
        self.assertEqual(sql.parse_datestr('2024-03-15'), ['2024', '03', '15'])

    def test_invalid_date_raises_value_error(self):
        with self.assertRaises(ValueError):
            sql.parse_datestr('not-a-date')

    def test_empty_string_raises_value_error(self):
        with self.assertRaises(ValueError):
            sql.parse_datestr('')


# ---------------------------------------------------------------------------
# CODE-7: load_result opens CSV file twice
# ---------------------------------------------------------------------------
class TestCode7LoadResultSingleOpen(unittest.TestCase):

    def test_csv_opened_only_once(self):
        import tempfile, csv as csv_mod
        from unittest.mock import patch, mock_open, MagicMock

        conn, cur = make_db()
        cur.execute("INSERT INTO Events VALUES (1, 'Test Race', 10.0, 0, 0, 0)")
        conn.commit()

        csv_content = 'name,age_group,club,hometown,time,pace\nAlice,F30,TC,Victoria,0:40:00,4:00\n'

        open_calls = []
        original_open = open

        def tracking_open(filename, *args, **kwargs):
            if str(filename).endswith('.csv'):
                open_calls.append(filename)
            return original_open(filename, *args, **kwargs)

        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False) as f:
            f.write(csv_content)
            tmpcsv = f.name

        with patch('builtins.open', side_effect=tracking_open):
            sql.load_result(conn, cur, tmpcsv, 'Test Race', '2024-01-01', 10.0, 0, 0, 0)

        self.assertEqual(len(open_calls), 1,
                         f"CSV file should be opened once, was opened {len(open_calls)} times")
        conn.close()


# ---------------------------------------------------------------------------
# CODE-4: get_workout_calculated_data returns WorkoutData dataclass
# ---------------------------------------------------------------------------
class TestCode4WorkoutDataclass(unittest.TestCase):

    def _wd(self, **kwargs):
        defaults = dict(date='2024-01-15', distance='10.0',
                        recovery='0:00:00', easy='1:00:00',
                        threshold='0:00:00', interval='0:00:00', repetition='0:00:00')
        defaults.update(kwargs)
        return sql.get_workout_calculated_data(**defaults)

    def test_returns_workout_data_instance(self):
        self.assertIsInstance(self._wd(), sql.WorkoutData)

    def test_isodate_str_format(self):
        wd = self._wd(date='2024-01-15')
        self.assertRegex(wd.isodate_str, r'^\d{4}-\d{2}-\d$')

    def test_p8020_all_easy(self):
        wd = self._wd(easy='1:00:00')
        self.assertAlmostEqual(wd.p8020, 100.0)

    def test_p8020_zero_time(self):
        wd = self._wd(easy='0:00:00')
        self.assertEqual(wd.p8020, 0.0)

    def test_jd_int_easy_only(self):
        # 60 mins easy at 0.2 pts/min = 12.0
        wd = self._wd(easy='1:00:00')
        self.assertAlmostEqual(wd.jd_int, 12.0)

    def test_add_workout_stores_correct_values(self):
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15', easy='1:00:00')
        cur.execute('SELECT p8020, jd_int FROM Log')
        row = cur.fetchone()
        self.assertAlmostEqual(row[0], 100.0)
        self.assertAlmostEqual(row[1], 12.0)
        conn.close()

    def test_change_workout_updates_calculated_fields(self):
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15', easy='1:00:00')
        cur.execute('SELECT id FROM Log')
        wo_id = cur.fetchone()[0]
        # Change to interval-only — jd_int should go up
        sql.change_workout(conn, cur, wo_id, '2024-01-15', 'Loc', 1,
                           'obj', 'notes', '10.0',
                           '0:00:00', '0:00:00', '0:00:00', '1:00:00', '0:00:00', 1)
        cur.execute('SELECT jd_int FROM Log WHERE id = ?', (wo_id,))
        jd = cur.fetchone()[0]
        self.assertAlmostEqual(jd, 60.0)  # 60 min interval at 1.0 pts/min
        conn.close()


# ---------------------------------------------------------------------------
# CODE-3: Consistent conn/cur/id signature ordering
# ---------------------------------------------------------------------------
class TestCode3SignatureOrdering(unittest.TestCase):

    def test_change_health_signature_order(self):
        """change_health(conn, cur, id, ...) — cur before id."""
        import inspect
        params = list(inspect.signature(sql.change_health).parameters)
        self.assertEqual(params[:3], ['conn', 'cur', 'id'])

    def test_change_workout_signature_order(self):
        """change_workout(conn, cur, id, ...) — conn/cur before id."""
        import inspect
        params = list(inspect.signature(sql.change_workout).parameters)
        self.assertEqual(params[:3], ['conn', 'cur', 'id'])

    def test_change_health_works_with_new_signature(self):
        conn, cur = make_db()
        sql.add_health(conn, cur, '2024-01-15', 70.0, 80.0, 85.0, 90.0, 95.0, 'note', 60)
        cur.execute('SELECT id FROM Health')
        hid = cur.fetchone()[0]
        sql.change_health(conn, cur, hid, '2024-01-16', 71.0, 80.0, 85.0, 90.0, 95.0, 'updated', 62)
        cur.execute('SELECT date, weight FROM Health WHERE id = ?', (hid,))
        row = cur.fetchone()
        self.assertEqual(row[0], '2024-01-16')
        self.assertAlmostEqual(row[1], 71.0)
        conn.close()

    def test_change_workout_works_with_new_signature(self):
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15', distance='10.0')
        cur.execute('SELECT id FROM Log')
        wo_id = cur.fetchone()[0]
        sql.change_workout(conn, cur, wo_id, '2024-01-16', 'NewLoc', 1,
                           'new obj', 'new notes', '12.0',
                           '0:00:00', '1:00:00', '0:00:00', '0:00:00', '0:00:00', 1)
        cur.execute('SELECT date, location, dist FROM Log WHERE id = ?', (wo_id,))
        row = cur.fetchone()
        self.assertEqual(row[0], '2024-01-16')
        self.assertEqual(row[1], 'NewLoc')
        self.assertAlmostEqual(row[2], 12.0)
        conn.close()


# ---------------------------------------------------------------------------
# CODE-5: Four duplicated stats functions collapsed via _build_stats_table
# ---------------------------------------------------------------------------
class TestCode5StatsDuplication(unittest.TestCase):

    def _many_workouts_db(self):
        from datetime import date, timedelta
        conn, cur = make_db()
        base = date(2023, 1, 2)
        for i in range(60):
            d = (base + timedelta(weeks=i)).strftime('%Y-%m-%d')
            add_wo(conn, cur, date=d)
        return conn, cur

    def test_get_log_52weeks_returns_52_rows(self):
        conn, cur = self._many_workouts_db()
        [intro, thead, tbody, summary] = sql.get_log_52weeks(conn, cur)
        self.assertEqual(len(tbody), 52)
        conn.close()

    def test_get_log_week_stats_returns_11_rows(self):
        conn, cur = self._many_workouts_db()
        [_, _, tbody, _] = sql.get_log_week_stats(conn, cur)
        self.assertEqual(len(tbody), 11)
        conn.close()

    def test_get_log_12months_returns_12_rows(self):
        conn, cur = self._many_workouts_db()
        [_, _, tbody, _] = sql.get_log_12months(conn, cur)
        self.assertEqual(len(tbody), 12)
        conn.close()

    def test_get_log_month_stats_returns_12_rows(self):
        conn, cur = self._many_workouts_db()
        [_, _, tbody, _] = sql.get_log_month_stats(conn, cur)
        self.assertEqual(len(tbody), 12)
        conn.close()

    def test_week_stats_first_row_label(self):
        conn, cur = self._many_workouts_db()
        [_, _, tbody, _] = sql.get_log_week_stats(conn, cur)
        self.assertEqual(tbody[0][0], 'This Week')
        conn.close()

    def test_month_stats_first_row_label(self):
        conn, cur = self._many_workouts_db()
        [_, _, tbody, _] = sql.get_log_month_stats(conn, cur)
        self.assertEqual(tbody[0][0], 'This Month')
        conn.close()

    def test_build_stats_table_helper_exists(self):
        self.assertTrue(callable(sql._build_stats_table))


# ---------------------------------------------------------------------------
# CODE-8: get_log_sums_for_all_weeks/months returned None silently
# ---------------------------------------------------------------------------
class TestCode8SilentNoneReturn(unittest.TestCase):

    def _make_db_with_workouts(self):
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15')
        add_wo(conn, cur, date='2024-06-10')
        return conn, cur

    def test_get_log_sums_for_all_weeks_returns_list(self):
        conn, cur = self._make_db_with_workouts()
        result = sql.get_log_sums_for_all_weeks(conn, cur)
        self.assertIsNotNone(result)
        self.assertIsInstance(result, list)
        conn.close()

    def test_get_log_sums_for_all_weeks_row_count(self):
        conn, cur = self._make_db_with_workouts()
        result = sql.get_log_sums_for_all_weeks(conn, cur)
        # Two workouts in different ISO weeks
        self.assertEqual(len(result), 2)
        conn.close()

    def test_get_log_sums_for_all_months_returns_list(self):
        conn, cur = self._make_db_with_workouts()
        result = sql.get_log_sums_for_all_months(conn, cur)
        self.assertIsNotNone(result)
        self.assertIsInstance(result, list)
        conn.close()

    def test_get_log_sums_for_all_months_row_count(self):
        conn, cur = self._make_db_with_workouts()
        result = sql.get_log_sums_for_all_months(conn, cur)
        # Jan and Jun are different months
        self.assertEqual(len(result), 2)
        conn.close()


# ---------------------------------------------------------------------------
# LOW-5: print() replaced with logging
# ---------------------------------------------------------------------------
class TestLow5Logging(unittest.TestCase):

    def test_no_bare_print_in_sql_py(self):
        import ast, pathlib
        src = pathlib.Path('sql.py').read_text()
        tree = ast.parse(src)
        prints = [
            node for node in ast.walk(tree)
            if isinstance(node, ast.Call)
            and isinstance(node.func, ast.Name)
            and node.func.id == 'print'
        ]
        self.assertEqual(prints, [], f"Found {len(prints)} bare print() call(s) in sql.py")

    def test_no_bare_print_in_races_py(self):
        import ast, pathlib
        src = pathlib.Path('races.py').read_text()
        tree = ast.parse(src)
        prints = [
            node for node in ast.walk(tree)
            if isinstance(node, ast.Call)
            and isinstance(node.func, ast.Name)
            and node.func.id == 'print'
        ]
        self.assertEqual(prints, [], f"Found {len(prints)} bare print() call(s) in races.py")

    def test_sql_uses_logger(self):
        import pathlib
        src = pathlib.Path('sql.py').read_text()
        self.assertIn('logger = logging.getLogger(__name__)', src)

    def test_add_workout_emits_info_log(self):
        import logging
        conn, cur = make_db()
        with self.assertLogs('sql', level='INFO') as cm:
            add_wo(conn, cur)
        self.assertTrue(any('Adding workout' in m for m in cm.output))
        conn.close()

    def test_error_path_emits_error_log(self):
        import logging
        conn, cur = make_db()
        sql.add_health(conn, cur, '2024-01-15', 70.0, 80.0, 85.0, 90.0, 95.0, 'note', 60)
        # Duplicate insert should trigger logger.error
        with self.assertLogs('sql', level='ERROR') as cm:
            sql.add_health(conn, cur, '2024-01-15', 71.0, 80.0, 85.0, 90.0, 95.0, 'dup', 60)
        self.assertTrue(any('Error' in m or 'error' in m for m in cm.output))
        conn.close()


# ---------------------------------------------------------------------------
# LOW-6: Nested <form> tags fixed in AddWorkout.html
# ---------------------------------------------------------------------------
class TestLow6NestedForms(unittest.TestCase):

    def test_only_one_form_element(self):
        import pathlib, re
        src = pathlib.Path('templates/AddWorkout.html').read_text()
        opening_forms = re.findall(r'<form[\s>]', src, re.IGNORECASE)
        closing_forms = re.findall(r'</form>', src, re.IGNORECASE)
        self.assertEqual(len(opening_forms), 1, "Should have exactly one opening <form>")
        self.assertEqual(len(closing_forms), 1, "Should have exactly one closing </form>")

    def test_form_has_post_method(self):
        import pathlib
        src = pathlib.Path('templates/AddWorkout.html').read_text()
        self.assertIn('method="POST"', src)


# ---------------------------------------------------------------------------
# LOW-7: Unquoted value attributes fixed in AddWorkout.html
# ---------------------------------------------------------------------------
class TestLow7QuotedAttributes(unittest.TestCase):

    def test_no_unquoted_value_attributes(self):
        import pathlib, re
        src = pathlib.Path('templates/AddWorkout.html').read_text()
        # Match value = {{ ... }} without surrounding quotes
        unquoted = re.findall(r'value\s*=\s*\{\{[^"\']*\}\}', src)
        self.assertEqual(unquoted, [],
                         f"Found unquoted value attributes: {unquoted}")

    def test_value_attributes_are_quoted(self):
        import pathlib, re
        src = pathlib.Path('templates/AddWorkout.html').read_text()
        quoted = re.findall(r'value\s*=\s*"[^"]*\{\{[^}]*\}\}[^"]*"', src)
        self.assertGreater(len(quoted), 0, "Expected quoted value=\"{{ ... }}\" attributes")


# ---------------------------------------------------------------------------
# VAL-1: _validate_workout
# ---------------------------------------------------------------------------
class TestValidateWorkout(unittest.TestCase):

    def _call(self, date='2024-01-15', distance='10.0', recovery='0:00:00',
              easy='1:00:00', threshold='0:00:00', interval='0:00:00', repetition='0:00:00'):
        return _validate_workout(date, distance, recovery, easy, threshold, interval, repetition)

    def test_valid_inputs_returns_empty(self):
        self.assertEqual(self._call(), [])

    def test_bad_date_format_reported(self):
        errors = self._call(date='15-01-2024')
        self.assertTrue(any('Date' in e for e in errors))

    def test_zero_distance_reported(self):
        errors = self._call(distance='0.0')
        self.assertTrue(any('Distance' in e for e in errors))

    def test_negative_distance_reported(self):
        errors = self._call(distance='-5')
        self.assertTrue(any('Distance' in e for e in errors))

    def test_non_numeric_distance_reported(self):
        errors = self._call(distance='abc')
        self.assertTrue(any('Distance' in e for e in errors))

    def test_bad_recovery_time_reported(self):
        errors = self._call(recovery='60:00')
        self.assertTrue(any('Recovery' in e for e in errors))

    def test_invalid_minutes_in_easy_reported(self):
        errors = self._call(easy='1:60:00')
        self.assertTrue(any('Easy' in e for e in errors))

    def test_multiple_bad_fields_all_reported(self):
        errors = self._call(date='bad', distance='-1',
                            recovery='bad', easy='bad',
                            threshold='bad', interval='bad', repetition='bad')
        self.assertGreater(len(errors), 3)


# ---------------------------------------------------------------------------
# VAL-2: _validate_health
# ---------------------------------------------------------------------------
class TestValidateHealth(unittest.TestCase):

    def _call(self, date='2024-01-15', weight='70.0', waist='80.0',
              waist_bb='85.0', hips='90.0', chest='95.0', hr='60'):
        return _validate_health(date, weight, waist, waist_bb, hips, chest, hr)

    def test_valid_inputs_returns_empty(self):
        self.assertEqual(self._call(), [])

    def test_bad_date_reported(self):
        errors = self._call(date='not-a-date')
        self.assertTrue(any('Date' in e for e in errors))

    def test_negative_weight_reported(self):
        errors = self._call(weight='-1.0')
        self.assertTrue(any('Weight' in e for e in errors))

    def test_non_numeric_waist_reported(self):
        errors = self._call(waist='abc')
        self.assertTrue(any('Waist' in e for e in errors))

    def test_zero_hr_reported(self):
        errors = self._call(hr='0')
        self.assertTrue(any('Heart rate' in e for e in errors))

    def test_negative_hr_reported(self):
        errors = self._call(hr='-10')
        self.assertTrue(any('Heart rate' in e for e in errors))

    def test_non_numeric_hr_reported(self):
        errors = self._call(hr='fast')
        self.assertTrue(any('Heart rate' in e for e in errors))


# ---------------------------------------------------------------------------
# VAL-3: _validate_shoe
# ---------------------------------------------------------------------------
class TestValidateShoe(unittest.TestCase):

    def test_valid_inputs_returns_empty(self):
        self.assertEqual(_validate_shoe('Nike', 'Nike Air Zoom'), [])

    def test_empty_short_name_reported(self):
        self.assertTrue(len(_validate_shoe('', 'Nike Air Zoom')) > 0)

    def test_whitespace_only_short_name_reported(self):
        self.assertTrue(len(_validate_shoe('   ', 'Nike Air Zoom')) > 0)

    def test_empty_long_name_reported(self):
        self.assertTrue(len(_validate_shoe('Nike', '')) > 0)

    def test_both_empty_gives_two_errors(self):
        self.assertEqual(len(_validate_shoe('', '')), 2)


# ---------------------------------------------------------------------------
# SQL-RET: retire_shoe and get_all_shoes
# ---------------------------------------------------------------------------
class TestRetireShoe(unittest.TestCase):

    def test_retire_active_shoe(self):
        conn, cur = make_db()
        sql.retire_shoe(conn, cur, 1)
        cur.execute('SELECT retired FROM Shoes WHERE id = 1')
        self.assertEqual(cur.fetchone()[0], 1)
        conn.close()

    def test_unretire_retired_shoe(self):
        conn, cur = make_db()
        cur.execute('UPDATE Shoes SET retired = 1 WHERE id = 1')
        conn.commit()
        sql.retire_shoe(conn, cur, 1)
        cur.execute('SELECT retired FROM Shoes WHERE id = 1')
        self.assertEqual(cur.fetchone()[0], 0)
        conn.close()

    def test_retire_nonexistent_shoe_does_not_crash(self):
        conn, cur = make_db()
        try:
            sql.retire_shoe(conn, cur, 9999)
        except Exception as e:
            self.fail(f'retire_shoe raised on missing id: {e}')
        conn.close()

    def test_get_all_shoes_includes_all_retired_states(self):
        conn, cur = make_db()
        cur.execute("INSERT INTO Shoes VALUES (2, 'OldShoe', 'Old Shoe Long', 1)")
        conn.commit()
        _, _, tbody, _ = sql.get_all_shoes(cur)
        short_names = [str(row[0]) for row in tbody]
        self.assertIn('TestShoe', short_names)
        self.assertIn('OldShoe', short_names)
        conn.close()

    def test_get_all_shoes_active_shows_retire_link(self):
        conn, cur = make_db()
        _, _, tbody, _ = sql.get_all_shoes(cur)
        action = str(tbody[0][5])
        self.assertIn('/RetireShoe/', action)
        self.assertIn('Retire', action)
        conn.close()

    def test_get_all_shoes_retired_shows_unretire_link(self):
        conn, cur = make_db()
        cur.execute('UPDATE Shoes SET retired = 1 WHERE id = 1')
        conn.commit()
        _, _, tbody, _ = sql.get_all_shoes(cur)
        action = str(tbody[0][5])
        self.assertIn('Unretire', action)
        conn.close()

    def test_get_all_shoes_summary_counts_all(self):
        conn, cur = make_db()
        cur.execute("INSERT INTO Shoes VALUES (2, 'OldShoe', 'Old Shoe Long', 1)")
        conn.commit()
        _, _, _, summary = sql.get_all_shoes(cur)
        self.assertIn('2', summary)
        conn.close()


# ---------------------------------------------------------------------------
# SQL-DICT: get_running_dict structure and content
# ---------------------------------------------------------------------------
class TestGetRunningDict(unittest.TestCase):

    def test_empty_db_returns_empty_dict(self):
        conn, cur = make_db()
        result = sql.get_running_dict(conn, cur)
        self.assertEqual(result, {})
        conn.close()

    def test_keys_are_sequential_from_one(self):
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15')
        add_wo(conn, cur, date='2024-01-16')
        result = sql.get_running_dict(conn, cur)
        self.assertEqual(sorted(result.keys()), [1, 2])
        conn.close()

    def test_entry_has_required_keys(self):
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15')
        entry = sql.get_running_dict(conn, cur)[1]
        for key in ('id', 'date', 'weekday', 'description', 'objective',
                    'notes', 'distance', 'time_run', 'pace', 'shoes'):
            self.assertIn(key, entry, f'Missing key: {key}')
        conn.close()

    def test_weekday_is_correct(self):
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15')   # Monday
        add_wo(conn, cur, date='2024-01-20')   # Saturday
        result = sql.get_running_dict(conn, cur)
        self.assertEqual(result[1]['weekday'], 'Monday')
        self.assertEqual(result[2]['weekday'], 'Saturday')
        conn.close()

    def test_ordered_by_date(self):
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-06-01')
        add_wo(conn, cur, date='2024-01-15')
        result = sql.get_running_dict(conn, cur)
        dates = [result[i]['date'].strftime('%Y-%m-%d') for i in sorted(result)]
        self.assertEqual(dates, ['2024-01-15', '2024-06-01'])
        conn.close()

    def test_distance_matches_inserted_value(self):
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15', distance='15.5')
        result = sql.get_running_dict(conn, cur)
        self.assertAlmostEqual(result[1]['distance'], 15.5)
        conn.close()


# ---------------------------------------------------------------------------
# ROUTE: Flask HTTP layer
# ---------------------------------------------------------------------------
def _make_temp_db_file() -> str:
    """Create a seeded SQLite temp file and return its path."""
    import tempfile
    fd, path = tempfile.mkstemp(suffix='.sqlite')
    os.close(fd)
    conn = sqlite3.connect(path)
    cur = conn.cursor()
    cur.execute('CREATE TABLE wo_type (id INTEGER PRIMARY KEY, type TEXT)')
    cur.execute('CREATE TABLE Shoes (id INTEGER PRIMARY KEY, shortName TEXT, longName TEXT, retired INTEGER)')
    cur.execute('''CREATE TABLE Log (
        id INTEGER PRIMARY KEY, date TEXT, location TEXT, wo_type INTEGER,
        objective TEXT, notes TEXT, dist REAL, time TEXT, time_secs REAL,
        pace TEXT, pace_secs REAL, recovery TEXT, recovery_secs REAL,
        easy TEXT, easy_secs REAL, threshold TEXT, threshold_secs REAL,
        interval TEXT, interval_secs REAL, repetition TEXT, repetition_secs REAL,
        p8020 REAL, jd_int REAL, shoeID INTEGER, isodate TEXT
    )''')
    cur.execute('''CREATE TABLE Events (
        id INTEGER PRIMARY KEY, eventname TEXT, dist REAL,
        min_elev INTEGER, max_elev INTEGER, gain_elev INTEGER, UNIQUE (eventname)
    )''')
    cur.execute('CREATE TABLE Races (id INTEGER PRIMARY KEY, eventID INTEGER, date TEXT, UNIQUE (eventID, date))')
    cur.execute('''CREATE TABLE Athletes (
        id INTEGER PRIMARY KEY, name TEXT, age_group TEXT, club TEXT, hometown TEXT,
        UNIQUE (name, hometown)
    )''')
    cur.execute('CREATE TABLE RaceTimes (id INTEGER PRIMARY KEY, athleteID INTEGER, raceID INTEGER, str_time TEXT, sec_time INTEGER, pace TEXT)')
    cur.execute('''CREATE TABLE Health (
        id INTEGER PRIMARY KEY, date TEXT, weight REAL, smallWaist REAL, bbWaist REAL,
        hip REAL, chest REAL, notes TEXT, HR REAL, UNIQUE (date)
    )''')
    cur.execute("INSERT INTO wo_type VALUES (1, 'Run')")
    cur.execute("INSERT INTO wo_type VALUES (2, 'Ride')")
    cur.execute("INSERT INTO Shoes VALUES (1, 'TestShoe', 'Test Shoe Long', 0)")
    conn.commit()
    conn.close()
    return path


class TestFlaskRoutes(unittest.TestCase):
    """Test the Flask HTTP layer using the test client against a temp SQLite file."""

    def setUp(self):
        self.db_path = _make_temp_db_file()
        self._patcher = patch('races.DATABASE', self.db_path)
        self._patcher.start()
        app.config['TESTING'] = True
        self.client = app.test_client()

    def tearDown(self):
        self._patcher.stop()
        os.unlink(self.db_path)

    # --- GET routes ---

    def test_homepage_returns_200(self):
        self.assertEqual(self.client.get('/').status_code, 200)

    def test_list_workouts_returns_200(self):
        self.assertEqual(self.client.get('/ListWorkouts').status_code, 200)

    def test_add_workout_get_returns_200(self):
        self.assertEqual(self.client.get('/AddWorkout').status_code, 200)

    def test_list_shoes_returns_200(self):
        self.assertEqual(self.client.get('/ListShoes').status_code, 200)

    def test_add_shoes_get_returns_200(self):
        self.assertEqual(self.client.get('/AddShoes').status_code, 200)

    def test_manage_shoes_returns_200(self):
        self.assertEqual(self.client.get('/ManageShoes').status_code, 200)

    def test_retire_shoe_returns_200(self):
        self.assertEqual(self.client.get('/RetireShoe/1').status_code, 200)

    def test_list_health_returns_200(self):
        self.assertEqual(self.client.get('/ListHealth').status_code, 200)

    def test_add_health_get_returns_200(self):
        self.assertEqual(self.client.get('/AddHealth').status_code, 200)

    def test_list_athletes_returns_200(self):
        self.assertEqual(self.client.get('/ListAthletes').status_code, 200)

    def test_compare_get_returns_200(self):
        self.assertEqual(self.client.get('/Compare').status_code, 200)

    def test_workout_52weeks_returns_200(self):
        self.assertEqual(self.client.get('/Workout52Weeks').status_code, 200)

    def test_workout_week_stats_returns_200(self):
        self.assertEqual(self.client.get('/WorkoutWeekStats').status_code, 200)

    # --- POST /AddWorkout ---

    _VALID_WORKOUT = dict(
        date='2024-01-15', location='TrackA', wo_type='Run',
        objective='Test', notes='Test notes', distance='10.0',
        recovery='0:00:00', easy='1:00:00', threshold='0:00:00',
        interval='0:00:00', repetition='0:00:00', shoes='TestShoe',
    )

    def test_add_workout_post_valid_shows_workout_list(self):
        r = self.client.post('/AddWorkout', data=self._VALID_WORKOUT, follow_redirects=True)
        self.assertEqual(r.status_code, 200)
        self.assertIn(b'Workouts', r.data)

    def test_add_workout_post_invalid_date_shows_error(self):
        r = self.client.post('/AddWorkout', data={**self._VALID_WORKOUT, 'date': 'bad-date'})
        self.assertEqual(r.status_code, 200)
        self.assertIn(b'Date must be', r.data)

    def test_add_workout_post_invalid_distance_shows_error(self):
        r = self.client.post('/AddWorkout', data={**self._VALID_WORKOUT, 'distance': '-5'})
        self.assertEqual(r.status_code, 200)
        self.assertIn(b'Distance', r.data)

    def test_add_workout_post_preserves_location_on_error(self):
        r = self.client.post('/AddWorkout', data={**self._VALID_WORKOUT, 'date': 'bad', 'location': 'MyTrack'})
        self.assertIn(b'MyTrack', r.data)

    def test_add_workout_post_saves_to_db(self):
        self.client.post('/AddWorkout', data=self._VALID_WORKOUT)
        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()
        cur.execute('SELECT date, location FROM Log')
        row = cur.fetchone()
        conn.close()
        self.assertIsNotNone(row)
        self.assertEqual(row[0], '2024-01-15')
        self.assertEqual(row[1], 'TrackA')

    # --- POST /AddHealth ---

    _VALID_HEALTH = dict(
        date='2024-01-15', weight='70.0', waist='80.0', waist_bb='85.0',
        hips='90.0', chest='95.0', HR='60', notes='test',
    )

    def test_add_health_post_valid_shows_health_list(self):
        r = self.client.post('/AddHealth', data=self._VALID_HEALTH, follow_redirects=True)
        self.assertEqual(r.status_code, 200)
        self.assertIn(b'Health', r.data)

    def test_add_health_post_invalid_date_shows_error(self):
        r = self.client.post('/AddHealth', data={**self._VALID_HEALTH, 'date': 'bad-date'})
        self.assertIn(b'Date must be', r.data)

    def test_add_health_post_zero_hr_shows_error(self):
        r = self.client.post('/AddHealth', data={**self._VALID_HEALTH, 'HR': '0'})
        self.assertIn(b'Heart rate', r.data)

    def test_add_health_post_saves_to_db(self):
        self.client.post('/AddHealth', data=self._VALID_HEALTH)
        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()
        cur.execute('SELECT date, weight FROM Health')
        row = cur.fetchone()
        conn.close()
        self.assertIsNotNone(row)
        self.assertEqual(row[0], '2024-01-15')

    # --- POST /AddShoes ---

    def test_add_shoes_post_valid_saves_to_db(self):
        self.client.post('/AddShoes', data={'shortName': 'NewShoe', 'longName': 'New Shoe Long'})
        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()
        cur.execute("SELECT shortName FROM Shoes WHERE shortName = 'NewShoe'")
        self.assertIsNotNone(cur.fetchone())
        conn.close()

    def test_add_shoes_post_empty_short_name_shows_error(self):
        r = self.client.post('/AddShoes', data={'shortName': '', 'longName': 'New Shoe Long'})
        self.assertEqual(r.status_code, 200)
        self.assertIn(b'Short name is required', r.data)

    def test_add_shoes_post_empty_names_does_not_insert(self):
        self.client.post('/AddShoes', data={'shortName': '', 'longName': ''})
        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()
        cur.execute('SELECT COUNT(*) FROM Shoes')
        count = cur.fetchone()[0]
        conn.close()
        self.assertEqual(count, 1)  # only the seed shoe

    # --- /RetireShoe toggles DB state ---

    def test_retire_shoe_toggles_retired_flag_in_db(self):
        self.client.get('/RetireShoe/1')
        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()
        cur.execute('SELECT retired FROM Shoes WHERE id = 1')
        self.assertEqual(cur.fetchone()[0], 1)
        conn.close()

    def test_list_workouts_search_returns_200(self):
        self.assertEqual(self.client.get('/ListWorkouts?q=TrackA').status_code, 200)

    def test_list_workouts_search_filters_results(self):
        self.client.post('/AddWorkout', data=self._VALID_WORKOUT, follow_redirects=True)
        r = self.client.get('/ListWorkouts?q=TrackA')
        self.assertIn(b'TrackA', r.data)

    def test_list_workouts_phrase_match(self):
        self.client.post('/AddWorkout', data=self._VALID_WORKOUT, follow_redirects=True)
        # Quoted phrase must match as a substring
        r = self.client.get('/ListWorkouts?q=%22Test+notes%22')
        self.assertIn(b'TrackA', r.data)

    def test_list_workouts_phrase_no_match_when_words_split(self):
        self.client.post('/AddWorkout', data=self._VALID_WORKOUT, follow_redirects=True)
        # "notes Test" is not a contiguous substring even though both words exist
        r = self.client.get('/ListWorkouts?q=%22notes+Test%22')
        self.assertIn(b'0 workout', r.data)

    def test_list_workouts_multiword_and_match(self):
        self.client.post('/AddWorkout', data=self._VALID_WORKOUT, follow_redirects=True)
        # Both words present in different fields — should match
        r = self.client.get('/ListWorkouts?q=TrackA+Test')
        self.assertIn(b'TrackA', r.data)

    def test_list_workouts_multiword_and_no_match(self):
        self.client.post('/AddWorkout', data=self._VALID_WORKOUT, follow_redirects=True)
        # Second word not in any field — should return zero
        r = self.client.get('/ListWorkouts?q=TrackA+xyzzy_no_match')
        self.assertIn(b'0 workout', r.data)

    def test_list_workouts_search_no_match_shows_zero(self):
        r = self.client.get('/ListWorkouts?q=xyzzy_no_match')
        self.assertIn(b'0 workout', r.data)

    def test_list_health_search_returns_200(self):
        self.assertEqual(self.client.get('/ListHealth?q=2024').status_code, 200)

    def test_list_workouts_pagination_param(self):
        self.assertEqual(self.client.get('/ListWorkouts?page=1').status_code, 200)

    def test_list_workouts_bad_page_param_defaults_to_1(self):
        self.assertEqual(self.client.get('/ListWorkouts?page=abc').status_code, 200)

    def test_retire_shoe_twice_restores_active(self):
        self.client.get('/RetireShoe/1')
        self.client.get('/RetireShoe/1')
        conn = sqlite3.connect(self.db_path)
        cur = conn.cursor()
        cur.execute('SELECT retired FROM Shoes WHERE id = 1')
        self.assertEqual(cur.fetchone()[0], 0)
        conn.close()


# ---------------------------------------------------------------------------
# GARMIN-1: garmin.format_lap_notes
# ---------------------------------------------------------------------------
class TestFormatLapNotes(unittest.TestCase):

    _LAPS = [
        {'averageSpeed': 3.984, 'averageHR': 135, 'distance': 1000.0},  # ~4:11/km
        {'averageSpeed': 3.968, 'averageHR': 141, 'distance': 1000.0},  # ~4:12/km
        {'averageSpeed': 3.953, 'averageHR': 144, 'distance': 1000.0},  # ~4:13/km
    ]

    def test_header_line(self):
        import garmin as g
        notes = g.format_lap_notes(self._LAPS)
        self.assertIn('Rep / Dist / Pace / AHR / RPE / Notes', notes)

    def test_correct_row_count(self):
        import garmin as g
        notes = g.format_lap_notes(self._LAPS)
        lines = notes.strip().splitlines()
        # header + 3 lap rows + blank line + total line = 6
        self.assertEqual(len(lines), 6)

    def test_rpe_column_is_blank(self):
        import garmin as g
        notes = g.format_lap_notes(self._LAPS)
        lap_lines = [l for l in notes.splitlines() if l and l[0].isdigit()]
        for line in lap_lines:
            parts = line.split(' / ')
            self.assertEqual(parts[4].strip(), '', f'RPE column should be blank: {line}')

    def test_total_line_present(self):
        import garmin as g
        notes = g.format_lap_notes(self._LAPS)
        self.assertIn('Total distance was 3.00 km', notes)

    def test_empty_laps_returns_total_only(self):
        import garmin as g
        notes = g.format_lap_notes([])
        self.assertIn('Total distance was 0.00 km', notes)

    def test_zero_speed_shows_placeholder(self):
        import garmin as g
        laps = [{'averageSpeed': 0, 'averageHR': 140, 'distance': 1000.0}]
        notes = g.format_lap_notes(laps)
        self.assertIn('--:--', notes)

    def test_started_at_line_when_time_provided(self):
        import garmin as g
        notes = g.format_lap_notes(self._LAPS, start_time='07:30')
        self.assertTrue(notes.startswith('Started at 07:30.'))

    def test_no_started_at_line_when_no_time(self):
        import garmin as g
        notes = g.format_lap_notes(self._LAPS)
        self.assertFalse(notes.startswith('Started at'))

    def test_blank_line_before_total(self):
        import garmin as g
        notes = g.format_lap_notes(self._LAPS)
        lines = notes.splitlines()
        total_idx = next(i for i, l in enumerate(lines) if l.startswith('Total'))
        self.assertEqual(lines[total_idx - 1], '')


# ---------------------------------------------------------------------------
# GARMIN-2: garmin.find_unlogged
# ---------------------------------------------------------------------------
class TestFindUnlogged(unittest.TestCase):

    _ACTIVITIES = [
        {'activityId': 1, 'startTimeLocal': '2024-01-15 10:00:00',
         'distance': 10000.0, 'activityName': 'Morning Run',
         'activityType': {'typeKey': 'running'}, 'duration': 3600},
        {'activityId': 2, 'startTimeLocal': '2024-01-20 09:00:00',
         'distance': 5000.0, 'activityName': 'Easy Run',
         'activityType': {'typeKey': 'running'}, 'duration': 1800},
    ]

    def test_already_logged_exact_match_excluded(self):
        import garmin as g
        conn, cur = make_db()
        add_wo(conn, cur, date='2024-01-15', distance='10.0')
        result = g.find_unlogged(cur, self._ACTIVITIES)
        ids = [a['activityId'] for a in result]
        self.assertNotIn(1, ids)
        conn.close()

    def test_unlogged_activity_included(self):
        import garmin as g
        conn, cur = make_db()
        result = g.find_unlogged(cur, self._ACTIVITIES)
        ids = [a['activityId'] for a in result]
        self.assertIn(1, ids)
        self.assertIn(2, ids)
        conn.close()

    def test_within_10_percent_excluded(self):
        import garmin as g
        conn, cur = make_db()
        # Log 9.5 km; Garmin 10.0 km — within 10%
        add_wo(conn, cur, date='2024-01-15', distance='9.5')
        result = g.find_unlogged(cur, self._ACTIVITIES)
        ids = [a['activityId'] for a in result]
        self.assertNotIn(1, ids)
        conn.close()

    def test_outside_10_percent_included(self):
        import garmin as g
        conn, cur = make_db()
        # Log 5.0 km; Garmin 10.0 km — outside 10%
        add_wo(conn, cur, date='2024-01-15', distance='5.0')
        result = g.find_unlogged(cur, self._ACTIVITIES)
        ids = [a['activityId'] for a in result]
        self.assertIn(1, ids)
        conn.close()


# ---------------------------------------------------------------------------
# GARMIN-3: garmin.map_to_form
# ---------------------------------------------------------------------------
class TestMapToForm(unittest.TestCase):

    _ACTIVITY = {
        'activityId': 42,
        'activityName': 'Track Session',
        'startTimeLocal': '2024-01-15 07:30:00',
        'distance': 8050.0,
        'duration': 2400.0,
        'activityType': {'typeKey': 'running'},
    }
    _LAPS = [
        {'averageSpeed': 4.0, 'averageHR': 142, 'distance': 1000.0},
    ]

    def test_date_extracted_correctly(self):
        import garmin as g
        form = g.map_to_form(self._ACTIVITY, self._LAPS)
        self.assertEqual(form['val_date'], '2024-01-15')

    def test_distance_in_km(self):
        import garmin as g
        form = g.map_to_form(self._ACTIVITY, self._LAPS)
        self.assertAlmostEqual(float(form['val_distance']), 8.05, places=1)

    def test_location_from_activity_name(self):
        import garmin as g
        form = g.map_to_form(self._ACTIVITY, self._LAPS)
        self.assertEqual(form['val_location'], 'Track Session')

    def test_running_maps_to_run(self):
        import garmin as g
        form = g.map_to_form(self._ACTIVITY, self._LAPS)
        self.assertEqual(form['val_wo_type'], 'Run')

    def test_cycling_maps_to_ride(self):
        import garmin as g
        act = {**self._ACTIVITY, 'activityType': {'typeKey': 'cycling'}}
        form = g.map_to_form(act, [])
        self.assertEqual(form['val_wo_type'], 'Ride')

    def test_notes_contains_lap_data(self):
        import garmin as g
        form = g.map_to_form(self._ACTIVITY, self._LAPS)
        self.assertIn('Rep / Dist / Pace / AHR / RPE / Notes', form['val_notes'])

    def test_time_zones_default_to_zero(self):
        import garmin as g
        form = g.map_to_form(self._ACTIVITY, self._LAPS)
        for field in ('val_recovery', 'val_easy', 'val_threshold',
                      'val_interval', 'val_repetition'):
            self.assertEqual(form[field], '0:00:00')


# ---------------------------------------------------------------------------
# GARMIN-4: Flask routes for Garmin integration
# ---------------------------------------------------------------------------
class TestGarminRoutes(unittest.TestCase):

    def setUp(self):
        self.db_path = _make_temp_db_file()
        self._db_patcher = patch('races.DATABASE', self.db_path)
        self._db_patcher.start()
        app.config['TESTING'] = True
        self.client = app.test_client()

    def tearDown(self):
        self._db_patcher.stop()
        os.unlink(self.db_path)

    def test_garmin_sync_missing_credentials_shows_error(self):
        with patch('garmin.get_client', side_effect=ValueError('Set GARMIN_EMAIL')):
            r = self.client.get('/GarminSync')
        self.assertEqual(r.status_code, 200)
        self.assertIn(b'Set GARMIN_EMAIL', r.data)

    def test_garmin_sync_connection_error_shows_error(self):
        with patch('garmin.get_client', side_effect=Exception('Network error')):
            r = self.client.get('/GarminSync')
        self.assertEqual(r.status_code, 200)
        self.assertIn(b'Garmin connection failed', r.data)

    def test_garmin_sync_no_unlogged_shows_success_message(self):
        with patch('garmin.get_client'), \
             patch('garmin.fetch_recent_activities', return_value=[]), \
             patch('garmin.find_unlogged', return_value=[]):
            r = self.client.get('/GarminSync')
        self.assertEqual(r.status_code, 200)
        self.assertIn(b'already in your log', r.data)

    def test_garmin_sync_shows_unlogged_activity(self):
        fake_act = {
            'activityId': 99,
            'startTimeLocal': '2024-06-01 08:00:00',
            'activityName': 'Morning Run',
            'activityType': {'typeKey': 'running'},
            'distance': 10000.0,
            'duration': 3600.0,
        }
        with patch('garmin.get_client'), \
             patch('garmin.fetch_recent_activities', return_value=[fake_act]), \
             patch('garmin.find_unlogged', return_value=[fake_act]):
            r = self.client.get('/GarminSync')
        self.assertIn(b'Morning Run', r.data)
        self.assertIn(b'99', r.data)

    def test_add_workout_from_garmin_returns_200(self):
        fake_form = {
            'val_date': '2024-06-01', 'val_location': 'Track',
            'val_objective': '', 'val_notes': 'Rep / Dist / Pace / AHR / RPE / Notes\n1 / 1.00 km / 4:11 / 135 /  \n\nTotal distance was 1.00 km in 4:11 minutes',
            'val_distance': '10.00', 'val_recovery': '0:00:00',
            'val_easy': '0:00:00', 'val_threshold': '0:00:00',
            'val_interval': '0:00:00', 'val_repetition': '0:00:00',
            'val_wo_type': 'Run',
        }
        with patch('garmin.get_client'), \
             patch('garmin.fetch_laps', return_value=[]), \
             patch('garmin.map_to_form', return_value=fake_form), \
             patch.object(__import__('garminconnect').Garmin, 'get_activity_details', return_value={}):
            r = self.client.get('/AddWorkoutFromGarmin/12345')
        self.assertEqual(r.status_code, 200)
        self.assertIn(b'Track', r.data)

    def test_add_workout_from_garmin_fetch_error_shows_error(self):
        with patch('garmin.get_client', side_effect=Exception('auth failed')):
            r = self.client.get('/AddWorkoutFromGarmin/12345')
        self.assertEqual(r.status_code, 200)
        self.assertIn(b'Could not load Garmin activity', r.data)


if __name__ == '__main__':
    unittest.main(verbosity=2)
